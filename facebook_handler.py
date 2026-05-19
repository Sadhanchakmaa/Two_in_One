# facebook_handler.py
import os
import openpyxl
from telegram import Update, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import ConversationHandler, MessageHandler, filters, CommandHandler, ContextTypes

from config import MAIN_MENU, BACK_FB_MENU, BASE_DIR

# স্টেটস
FB_SELECTING_OPTION = 10
FB_SETTING_PASSWORD = 11
FB_WAITING_UID = 12
FB_WAITING_PASSWORD = 13
FB_WAITING_PASSWORD_INPUT = 14
FB_WAITING_COOKIES = 15
FB_CONFIRM_SAVE = 16

CUSTOM_EMOJI = {
    "📱": "6210664181943771071",
    "📂": "5332586662629227075",
    "🌍": "6188045471118790922",
    "🔄": "5264727218734524899",
    "ℹ️": "5332679880599418983",
    "🔥": "6231224366483384495",
    "✅": "6230828091325818369",
    "🚀": "6147654280112248427",
    "📞": "5467539229468793355",
    "🔙": "6206505206197261313",
    "🏠": "5416041192905265756",
    "📎": "5377844313575150051",
    "✏️": "5395444784611480792",
    "✨": "6232983172770962452",
    "👑": "6233506609025261770",
    "💎": "6230898013393396414",
    "🎯": "5310278924616356636",
    "📊": "6233241506463883080",
    "🔒": "5296369303661067030",
    "📥": "5433811242135331842",
    "⚙️": "5895577117592128901",
    "🗑️": "6233541449799966206",
    "💾": "25431736674147114227",
    "📖": "5226512880362332956",
    "⏳": "6154603635981423575",
    "❌": "6232999605315837644",
    "📝": "5436353008076075648",
    "📌": "5436007417827578363",
    "🟢": "6113685078825505075",
    "🔴": "6233042091132329496",
    "🎲": "5472404950673791399",
    "📋": "5197269100878907942",
    "📄": "5873153278023307367",
    "🏷️": "6066473359493828085",
    "🔑": "5330115548900501467",
    "🍪": "6233183429916107004",
    "🆔": "5841276284155467413",
    "⚠️": "5436034420286960597",
    "👇": "6233121371933646400",
    "⏭️": "6233383992003927192",
    "🔐": "5472308992514464048"
}

def get_fb_file_path(user_id, option):
    if option == "1000X":
        return os.path.join(BASE_DIR, f"{user_id}_1000X.xlsx")
    else:
        return os.path.join(BASE_DIR, f"{user_id}_6155X.xlsx")


def get_fb_count(user_id, option):
    path = get_fb_file_path(user_id, option)
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        return wb.active.max_row
    return 0


def validate_fb_uid(uid, option):
    if option == "1000X":
        return uid.startswith("1000")
    else:
        return uid.startswith(("6155", "6156", "6157", "6158"))


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text(
        f'<tg-emoji emoji-id="{CUSTOM_EMOJI["✨"]}">✨</tg-emoji> Cancelled! Returning to Main Menu <tg-emoji emoji-id="{CUSTOM_EMOJI["🏠"]}">🏠</tg-emoji>',
        parse_mode='HTML',
        reply_markup=MAIN_MENU
    )
    return ConversationHandler.END


async def fb_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    context.user_data['fb_user_id'] = user_id
    count_1000 = get_fb_count(user_id, "1000X")
    count_6155 = get_fb_count(user_id, "6155X")
    
    fb_menu = ReplyKeyboardMarkup([
        [KeyboardButton(f"1000X({count_1000})",style="success", icon_custom_emoji_id=CUSTOM_EMOJI["📂"]), 
         KeyboardButton(f"6155X({count_6155})",style="primary", icon_custom_emoji_id=CUSTOM_EMOJI["📂"])],
        [KeyboardButton("Set Password",style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🔒"])],
        [KeyboardButton("Download",style="success", icon_custom_emoji_id=CUSTOM_EMOJI["📥"])],
        [KeyboardButton("Reset Data",style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🗑️"])],
        [KeyboardButton("BACK TO MAIN MENU",style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🔙"])]
    ], resize_keyboard=True)
    
    await update.message.reply_text(
        f'<tg-emoji emoji-id="{CUSTOM_EMOJI["👑"]}">👑</tg-emoji> <b>VIP FACEBOOK DATA MANAGER</b> <tg-emoji emoji-id="{CUSTOM_EMOJI["👑"]}">👑</tg-emoji>\n\n'
        f'<tg-emoji emoji-id="{CUSTOM_EMOJI["💎"]}">💎</tg-emoji> <b>Welcome to Team X Bot</b>\n'
        f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🎯"]}">🎯</tg-emoji> <b>System Ready</b>\n\n'
        f'<tg-emoji emoji-id="{CUSTOM_EMOJI["📊"]}">📊</tg-emoji> <b>Please select an option below:</b>',
        parse_mode='HTML',
        reply_markup=fb_menu
    )
    return FB_SELECTING_OPTION


async def fb_select_option(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    user_id = context.user_data.get('fb_user_id')
    
    if text == "BACK":
        count_1000 = get_fb_count(user_id, "1000X")
        count_6155 = get_fb_count(user_id, "6155X")
        fb_menu = ReplyKeyboardMarkup([
            [KeyboardButton(f"1000X({count_1000})", style="success", icon_custom_emoji_id=CUSTOM_EMOJI["📂"]), 
             KeyboardButton(f"6155X({count_6155})", style="primary", icon_custom_emoji_id=CUSTOM_EMOJI["📂"])],
            [KeyboardButton("Set Password", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🔒"]), 
             KeyboardButton("Download", style="success", icon_custom_emoji_id=CUSTOM_EMOJI["📥"])],
            [KeyboardButton("Reset Data", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🗑️"])],
            [KeyboardButton("BACK TO MAIN MENU", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🔙"])]
        ], resize_keyboard=True)
        await update.message.reply_text(
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["👑"]}">👑</tg-emoji> <b>FACEBOOK DATA MANAGER</b>\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["📊"]}">📊</tg-emoji> <b>Please select an option</b>',
            parse_mode='HTML', reply_markup=fb_menu
        )
        return FB_SELECTING_OPTION
    
    if text == "BACK TO MAIN MENU":
        await update.message.reply_text(
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🏠"]}">🏠</tg-emoji> <b>BACK TO MENU</b>',
            parse_mode='HTML', reply_markup=MAIN_MENU
        )
        return ConversationHandler.END
    
    # 1000X বাটন (UID যোগ করার জন্য)
    if text.startswith("1000X") and "File" not in text:
        context.user_data['fb_current_option'] = "1000X"
        path = get_fb_file_path(user_id, "1000X")
        if not os.path.exists(path):
            wb = openpyxl.Workbook()
            wb.save(path)
        await update.message.reply_text(
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["📱"]}">📱</tg-emoji> <b>ENTER FB ACCOUNT UID</b>\n\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["📝"]}">📝</tg-emoji> <b>Format:</b>\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🎯"]}">🎯</tg-emoji> Must start with 1000\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🎯"]}">🎯</tg-emoji> Example: 1000123456789\n\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🔙"]}">🔙</tg-emoji> BACK to cancel',
            parse_mode='HTML', reply_markup=BACK_FB_MENU
        )
        return FB_WAITING_UID
    
    # 6155X বাটন (UID যোগ করার জন্য)
    if text.startswith("6155X") and "File" not in text:
        context.user_data['fb_current_option'] = "6155X"
        path = get_fb_file_path(user_id, "6155X")
        if not os.path.exists(path):
            wb = openpyxl.Workbook()
            wb.save(path)
        await update.message.reply_text(
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["📱"]}">📱</tg-emoji> <b>ENTER FB ACCOUNT UID</b>\n\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["📝"]}">📝</tg-emoji> <b>Format:</b>\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🎯"]}">🎯</tg-emoji> Must start with:\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🎯"]}">🎯</tg-emoji> 6155, 6156, 6157, 6158\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🎯"]}">🎯</tg-emoji> Example: 6155123456789\n\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🔙"]}">🔙</tg-emoji> BACK to cancel',
            parse_mode='HTML', reply_markup=BACK_FB_MENU
        )
        return FB_WAITING_UID
    
    elif text == "Set Password":
        await update.message.reply_text(
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🔐"]}">🔐</tg-emoji> <b>SET GLOBAL PASSWORD</b>\n\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["✏️"]}">✏️</tg-emoji> Enter your password\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🎯"]}">🎯</tg-emoji> Type \'skip\' to continue without password\n\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🔙"]}">🔙</tg-emoji> BACK to cancel',
            parse_mode='HTML', reply_markup=BACK_FB_MENU
        )
        return FB_SETTING_PASSWORD
    
    elif text == "Reset Data":
        for opt in ["1000X", "6155X"]:
            path = get_fb_file_path(user_id, opt)
            if os.path.exists(path):
                os.remove(path)
        
        count_1000 = get_fb_count(user_id, "1000X")
        count_6155 = get_fb_count(user_id, "6155X")
        fb_menu = ReplyKeyboardMarkup([
            [KeyboardButton(f"1000X({count_1000})", style="success", icon_custom_emoji_id=CUSTOM_EMOJI["📂"]), 
             KeyboardButton(f"6155X({count_6155})", style="primary", icon_custom_emoji_id=CUSTOM_EMOJI["📂"])],
            [KeyboardButton("Set Password", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🔒"]), 
             KeyboardButton("Download", style="success", icon_custom_emoji_id=CUSTOM_EMOJI["📥"])],
            [KeyboardButton("Reset Data", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🗑️"])],
            [KeyboardButton("BACK TO MAIN MENU", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🔙"])]
        ], resize_keyboard=True)
        
        await update.message.reply_text(
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["✅"]}">✅</tg-emoji> <b>DATA CLEARED</b>\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🗑️"]}">🗑️</tg-emoji> All files have been deleted\n\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["👑"]}">👑</tg-emoji> <b>FACEBOOK DATA MANAGER</b>\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["📊"]}">📊</tg-emoji> <b>Please select an option</b>',
            parse_mode='HTML', reply_markup=fb_menu
        )
        return FB_SELECTING_OPTION
    
    elif text == "Download":
        download_menu = ReplyKeyboardMarkup([
            [KeyboardButton("1000X File", style="success", icon_custom_emoji_id=CUSTOM_EMOJI["📂"]), 
             KeyboardButton("6155X File", style="primary", icon_custom_emoji_id=CUSTOM_EMOJI["📂"])],
            [KeyboardButton("Both Files", style="success", icon_custom_emoji_id=CUSTOM_EMOJI["📂"])],
            [KeyboardButton("BACK", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🔙"])]
        ], resize_keyboard=True)
        await update.message.reply_text(
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["📥"]}">📥</tg-emoji> <b>DOWNLOAD CENTER</b>\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🎯"]}">🎯</tg-emoji> Select file to download:',
            parse_mode='HTML', reply_markup=download_menu
        )
        return FB_SELECTING_OPTION
    
    elif text == "1000X File":
        return await fb_download_file(update, context, "1000X")
    
    elif text == "6155X File":
        return await fb_download_file(update, context, "6155X")
    
    elif text == "Both Files":
        return await fb_download_both(update, context)
    
    else:
        await update.message.reply_text(
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["❌"]}">❌</tg-emoji> <b>INVALID OPTION</b>\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🎯"]}">🎯</tg-emoji> Please use the buttons',
            parse_mode='HTML'
        )
        return FB_SELECTING_OPTION


async def fb_download_file(update: Update, context: ContextTypes.DEFAULT_TYPE, option):
    user_id = context.user_data.get('fb_user_id')
    path = get_fb_file_path(user_id, option)
    
    if not os.path.exists(path):
        await update.message.reply_text(
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["❌"]}">❌</tg-emoji> <b>NO DATA FOUND</b>\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["📂"]}">📂</tg-emoji> {option} has no data\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["📥"]}">📥</tg-emoji> Add data first',
            parse_mode='HTML'
        )
        return FB_SELECTING_OPTION
    
    wb = openpyxl.load_workbook(path)
    total = wb.active.max_row
    
    with open(path, "rb") as f:
        await update.message.reply_document(
            document=f, 
            filename=f"{option}.xlsx",
            caption=f'<tg-emoji emoji-id="{CUSTOM_EMOJI["✅"]}">✅</tg-emoji> <b>FILE READY FOR DOWNLOAD</b>\n\n'
                    f'<tg-emoji emoji-id="{CUSTOM_EMOJI["📂"]}">📂</tg-emoji> <b>File:</b> {option}\n'
                    f'<tg-emoji emoji-id="{CUSTOM_EMOJI["📊"]}">📊</tg-emoji> <b>Total:</b> {total} entries',
            parse_mode='HTML'
        )
    
    return FB_SELECTING_OPTION


async def fb_download_both(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = context.user_data.get('fb_user_id')
    sent_count = 0
    
    for opt in ["1000X", "6155X"]:
        path = get_fb_file_path(user_id, opt)
        if os.path.exists(path):
            wb = openpyxl.load_workbook(path)
            total = wb.active.max_row
            with open(path, "rb") as f:
                await update.message.reply_document(
                    document=f, 
                    filename=f"{opt}.xlsx",
                    caption=f'<tg-emoji emoji-id="{CUSTOM_EMOJI["📂"]}">📂</tg-emoji> <b>{opt}</b>\n'
                            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["📊"]}">📊</tg-emoji> <b>{total} entries</b>',
                    parse_mode='HTML'
                )
            sent_count += 1
        else:
            await update.message.reply_text(
                f'<tg-emoji emoji-id="{CUSTOM_EMOJI["⚠️"]}">⚠️</tg-emoji> <b>NO DATA FOR {opt}</b>',
                parse_mode='HTML'
            )
    
    if sent_count == 0:
        await update.message.reply_text(
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["❌"]}">❌</tg-emoji> <b>NO FILES AVAILABLE</b>\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🎯"]}">🎯</tg-emoji> No data found in both 1000X and 6155X folders',
            parse_mode='HTML'
        )
    
    return FB_SELECTING_OPTION


async def fb_set_password(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    user_id = context.user_data.get('fb_user_id')
    
    if text == "BACK":
        count_1000 = get_fb_count(user_id, "1000X")
        count_6155 = get_fb_count(user_id, "6155X")
        fb_menu = ReplyKeyboardMarkup([
            [KeyboardButton(f"1000X({count_1000})", style="success", icon_custom_emoji_id=CUSTOM_EMOJI["📂"]), 
             KeyboardButton(f"6155X({count_6155})", style="primary", icon_custom_emoji_id=CUSTOM_EMOJI["📂"])],
            [KeyboardButton("Set Password", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🔒"])], 
             [KeyboardButton("Download", style="success", icon_custom_emoji_id=CUSTOM_EMOJI["📥"])],
            [KeyboardButton("Reset Data", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🗑️"])],
            [KeyboardButton("BACK TO MAIN MENU", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🔙"])]
        ], resize_keyboard=True)
        await update.message.reply_text(
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["👑"]}">👑</tg-emoji> <b>FACEBOOK DATA MANAGER</b>\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["📊"]}">📊</tg-emoji> <b>Please select an option</b>',
            parse_mode='HTML', reply_markup=fb_menu
        )
        return FB_SELECTING_OPTION
    
    if text.lower() == "skip":
        context.user_data['fb_global_password'] = None
        await update.message.reply_text(
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["⏭️"]}">⏭️</tg-emoji> <b>PASSWORD SKIPPED</b>\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🎯"]}">🎯</tg-emoji> No password will be set',
            parse_mode='HTML'
        )
    else:
        context.user_data['fb_global_password'] = text
        updated_count = 0
        for opt in ["1000X", "6155X"]:
            path = get_fb_file_path(user_id, opt)
            if os.path.exists(path):
                wb = openpyxl.load_workbook(path)
                ws = wb.active
                for row in range(1, ws.max_row + 1):
                    ws.cell(row=row, column=2, value=text)
                wb.save(path)
                updated_count += 1
        await update.message.reply_text(
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🔐"]}">🔐</tg-emoji> <b>PASSWORD SET SUCCESSFULLY</b>\n\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["📂"]}">📂</tg-emoji> <b>Files updated:</b> {updated_count}\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🔒"]}">🔒</tg-emoji> <b>Password:</b> <code>{text[:10]}...</code>',
            parse_mode='HTML'
        )
    
    count_1000 = get_fb_count(user_id, "1000X")
    count_6155 = get_fb_count(user_id, "6155X")
    fb_menu = ReplyKeyboardMarkup([
        [KeyboardButton(f"1000X({count_1000})", style="success", icon_custom_emoji_id=CUSTOM_EMOJI["📂"]), 
         KeyboardButton(f"6155X({count_6155})", style="primary", icon_custom_emoji_id=CUSTOM_EMOJI["📂"])],
        [KeyboardButton("Set Password", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🔒"])], 
         [KeyboardButton("Download", style="success", icon_custom_emoji_id=CUSTOM_EMOJI["📥"])],
        [KeyboardButton("Reset Data", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🗑️"])],
        [KeyboardButton("BACK TO MAIN MENU", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🔙"])]
    ], resize_keyboard=True)
    await update.message.reply_text(
        f'<tg-emoji emoji-id="{CUSTOM_EMOJI["👑"]}">👑</tg-emoji> <b>FACEBOOK DATA MANAGER</b>\n'
        f'<tg-emoji emoji-id="{CUSTOM_EMOJI["📊"]}">📊</tg-emoji> <b>Please select an option</b>',
        parse_mode='HTML', reply_markup=fb_menu
    )
    return FB_SELECTING_OPTION


async def fb_waiting_uid(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    current_option = context.user_data.get('fb_current_option')
    
    if text == "BACK":
        user_id = context.user_data.get('fb_user_id')
        count_1000 = get_fb_count(user_id, "1000X")
        count_6155 = get_fb_count(user_id, "6155X")
        fb_menu = ReplyKeyboardMarkup([
            [KeyboardButton(f"1000X({count_1000})", style="success", icon_custom_emoji_id=CUSTOM_EMOJI["📂"]), 
             KeyboardButton(f"6155X({count_6155})", style="primary", icon_custom_emoji_id=CUSTOM_EMOJI["📂"])],
            [KeyboardButton("Set Password", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🔒"])], 
             [KeyboardButton("Download", style="success", icon_custom_emoji_id=CUSTOM_EMOJI["📥"])],
            [KeyboardButton("Reset Data", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🗑️"])],
            [KeyboardButton("BACK TO MAIN MENU", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🔙"])]
        ], resize_keyboard=True)
        await update.message.reply_text(
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["👑"]}">👑</tg-emoji> <b>FACEBOOK DATA MANAGER</b>\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["📊"]}">📊</tg-emoji> <b>Please select an option</b>',
            parse_mode='HTML', reply_markup=fb_menu
        )
        return FB_SELECTING_OPTION
    
    if not validate_fb_uid(text, current_option):
        await update.message.reply_text(
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["❌"]}">❌</tg-emoji> <b>INVALID UID FOR {current_option}</b>\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["📝"]}">📝</tg-emoji> Please enter a valid UID:\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🔙"]}">🔙</tg-emoji> BACK to cancel',
            parse_mode='HTML', reply_markup=BACK_FB_MENU
        )
        return FB_WAITING_UID
    
    context.user_data['fb_current_uid'] = text
    
    pwd_menu = ReplyKeyboardMarkup([
        [KeyboardButton("Use Global", style="success", icon_custom_emoji_id=CUSTOM_EMOJI["🔐"]), 
         KeyboardButton("New Password", style="primary", icon_custom_emoji_id=CUSTOM_EMOJI["✏️"])],
        [KeyboardButton("Skip", style="success", icon_custom_emoji_id=CUSTOM_EMOJI["⏭️"])],
        [KeyboardButton("BACK", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🔙"])]
    ], resize_keyboard=True)
    
    await update.message.reply_text(
        f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🔐"]}">🔐</tg-emoji> <b>PASSWORD OPTIONS</b>\n'
        f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🎯"]}">🎯</tg-emoji> How would you like to set the password?',
        parse_mode='HTML', reply_markup=pwd_menu
    )
    return FB_WAITING_PASSWORD


async def fb_waiting_password(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    
    if text == "BACK":
        user_id = context.user_data.get('fb_user_id')
        count_1000 = get_fb_count(user_id, "1000X")
        count_6155 = get_fb_count(user_id, "6155X")
        fb_menu = ReplyKeyboardMarkup([
            [KeyboardButton(f"1000X({count_1000})", style="success", icon_custom_emoji_id=CUSTOM_EMOJI["📂"]), 
             KeyboardButton(f"6155X({count_6155})", style="primary", icon_custom_emoji_id=CUSTOM_EMOJI["📂"])],
            [KeyboardButton("Set Password", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🔒"])], 
             [KeyboardButton("Download", style="success", icon_custom_emoji_id=CUSTOM_EMOJI["📥"])],
            [KeyboardButton("Reset Data", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🗑️"])],
            [KeyboardButton("BACK TO MAIN MENU", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🔙"])]
        ], resize_keyboard=True)
        await update.message.reply_text(
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["👑"]}">👑</tg-emoji> <b>FACEBOOK DATA MANAGER</b>\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["📊"]}">📊</tg-emoji> <b>Please select an option</b>',
            parse_mode='HTML', reply_markup=fb_menu
        )
        return FB_SELECTING_OPTION
    
    if text == "Use Global":
        pwd = context.user_data.get('fb_global_password')
        if not pwd:
            await update.message.reply_text(
                f'<tg-emoji emoji-id="{CUSTOM_EMOJI["❌"]}">❌</tg-emoji> <b>NO GLOBAL PASSWORD</b>\n'
                f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🎯"]}">🎯</tg-emoji> Please set a password first using \'Set Password\'',
                parse_mode='HTML', reply_markup=BACK_FB_MENU
            )
            return FB_WAITING_PASSWORD
        context.user_data['fb_current_password'] = pwd
        await update.message.reply_text(
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🍪"]}">🍪</tg-emoji> <b>ENTER COOKIES</b>\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🎯"]}">🎯</tg-emoji> Please enter your Facebook account cookies\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🎯"]}">🎯</tg-emoji> (or type \'skip\')\n\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🔙"]}">🔙</tg-emoji> BACK to cancel',
            parse_mode='HTML', reply_markup=BACK_FB_MENU
        )
        return FB_WAITING_COOKIES
    
    elif text == "New Password":
        await update.message.reply_text(
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🔑"]}">🔑</tg-emoji> <b>ENTER NEW PASSWORD</b>\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🎯"]}">🎯</tg-emoji> Please enter your new password for this account\n\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🔙"]}">🔙</tg-emoji> BACK to cancel',
            parse_mode='HTML', reply_markup=BACK_FB_MENU
        )
        return FB_WAITING_PASSWORD_INPUT
    
    elif text == "Skip":
        context.user_data['fb_current_password'] = ""
        await update.message.reply_text(
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🍪"]}">🍪</tg-emoji> <b>ENTER COOKIES</b>\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🎯"]}">🎯</tg-emoji> Please enter your Facebook account cookies\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🎯"]}">🎯</tg-emoji> (or type \'skip\')\n\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🔙"]}">🔙</tg-emoji> BACK to cancel',
            parse_mode='HTML', reply_markup=BACK_FB_MENU
        )
        return FB_WAITING_COOKIES
    
    else:
        await update.message.reply_text(
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["❌"]}">❌</tg-emoji> <b>INVALID OPTION</b>\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🎯"]}">🎯</tg-emoji> Please use the buttons',
            parse_mode='HTML'
        )
        return FB_WAITING_PASSWORD


async def fb_waiting_password_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    
    if text == "BACK":
        user_id = context.user_data.get('fb_user_id')
        count_1000 = get_fb_count(user_id, "1000X")
        count_6155 = get_fb_count(user_id, "6155X")
        fb_menu = ReplyKeyboardMarkup([
            [KeyboardButton(f"1000X({count_1000})", style="success", icon_custom_emoji_id=CUSTOM_EMOJI["📂"]), 
             KeyboardButton(f"6155X({count_6155})", style="primary", icon_custom_emoji_id=CUSTOM_EMOJI["📂"])],
            [KeyboardButton("Set Password", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🔒"])], 
             [KeyboardButton("Download", style="success", icon_custom_emoji_id=CUSTOM_EMOJI["📥"])],
            [KeyboardButton("Reset Data", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🗑️"])],
            [KeyboardButton("BACK TO MAIN MENU", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🔙"])]
        ], resize_keyboard=True)
        await update.message.reply_text(
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["👑"]}">👑</tg-emoji> <b>FACEBOOK DATA MANAGER</b>\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["📊"]}">📊</tg-emoji> <b>Please select an option</b>',
            parse_mode='HTML', reply_markup=fb_menu
        )
        return FB_SELECTING_OPTION
    
    context.user_data['fb_current_password'] = text
    
    await update.message.reply_text(
        f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🍪"]}">🍪</tg-emoji> <b>ENTER COOKIES</b>\n'
        f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🎯"]}">🎯</tg-emoji> Please enter your Facebook account cookies\n'
        f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🎯"]}">🎯</tg-emoji> (or type \'skip\')\n\n'
        f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🔙"]}">🔙</tg-emoji> BACK to cancel',
        parse_mode='HTML', reply_markup=BACK_FB_MENU
    )
    return FB_WAITING_COOKIES


async def fb_waiting_cookies(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    
    if text == "BACK":
        user_id = context.user_data.get('fb_user_id')
        count_1000 = get_fb_count(user_id, "1000X")
        count_6155 = get_fb_count(user_id, "6155X")
        fb_menu = ReplyKeyboardMarkup([
            [KeyboardButton(f"1000X({count_1000})", style="success", icon_custom_emoji_id=CUSTOM_EMOJI["📂"]), 
             KeyboardButton(f"6155X ({count_6155})", style="primary", icon_custom_emoji_id=CUSTOM_EMOJI["📂"])],
            [KeyboardButton("Set Password", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🔒"])], 
            [KeyboardButton("Download", style="success", icon_custom_emoji_id=CUSTOM_EMOJI["📥"])],
            [KeyboardButton("Reset Data", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🗑️"])],
            [KeyboardButton("BACK TO MAIN MENU", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🔙"])]
        ], resize_keyboard=True)
        await update.message.reply_text(
            "👑 *FACEBOOK DATA MANAGER*\n"
            "📊 *Please select an option*",
            parse_mode="Markdown", reply_markup=fb_menu
        )
        return FB_SELECTING_OPTION
    
    cookies = "" if text.lower() == "skip" else text
    
    context.user_data['fb_temp_uid'] = context.user_data.get('fb_current_uid')
    context.user_data['fb_temp_password'] = context.user_data.get('fb_current_password', '')
    context.user_data['fb_temp_cookies'] = cookies
    context.user_data['fb_temp_option'] = context.user_data.get('fb_current_option')
    
    confirm_menu = ReplyKeyboardMarkup([
        [KeyboardButton("Yes", style="success"), 
         KeyboardButton("No", style="danger")],
        [KeyboardButton("BACK", style="danger")]
    ], resize_keyboard=True)
    
    pwd_display = context.user_data['fb_temp_password'][:15] if context.user_data['fb_temp_password'] else "Skip"
    cookies_display = cookies[:30] + "..." if len(cookies) > 30 else (cookies if cookies else "Skip")
    
    await update.message.reply_text(
        f"✨ *PREVIEW DATA* ✨\n\n"
        f"🆔 *UID:* `{context.user_data['fb_temp_uid']}`\n"
        f"🔐 *Password:* `{pwd_display}`\n"
        f"🍪 *Cookies:* `{cookies_display}`\n\n"
        f"💾 *Do You Save this Facebook Account  data?*",
        parse_mode="Markdown", reply_markup=confirm_menu
    )
    return FB_CONFIRM_SAVE


async def fb_confirm_save(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    user_id = context.user_data.get('fb_user_id')
    
    if text == "BACK":
        count_1000 = get_fb_count(user_id, "1000X")
        count_6155 = get_fb_count(user_id, "6155X")
        fb_menu = ReplyKeyboardMarkup([
            [KeyboardButton(f"1000X({count_1000})", style="success", icon_custom_emoji_id=CUSTOM_EMOJI["📂"]), 
             KeyboardButton(f"6155X({count_6155})", style="primary", icon_custom_emoji_id=CUSTOM_EMOJI["📂"])],
            [KeyboardButton("Set Password", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🔒"])],
            [KeyboardButton("Download", style="success", icon_custom_emoji_id=CUSTOM_EMOJI["📥"])],
            [KeyboardButton("Reset Data", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🗑️"])],
            [KeyboardButton("BACK TO MAIN MENU", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🔙"])]
        ], resize_keyboard=True)
        await update.message.reply_text(
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["👑"]}">👑</tg-emoji> <b>FACEBOOK DATA MANAGER</b>\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["📊"]}">📊</tg-emoji> <b>Please select an option</b>',
            parse_mode='HTML', reply_markup=fb_menu
        )
        return FB_SELECTING_OPTION
    
    if text == "Yes":
        uid = context.user_data.get('fb_temp_uid')
        pwd = context.user_data.get('fb_temp_password', '')
        ck = context.user_data.get('fb_temp_cookies', '')
        opt = context.user_data.get('fb_temp_option')
        
        path = get_fb_file_path(user_id, opt)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        
        row_num = None
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == uid:
                row_num = row
                break
        
        if row_num:
            ws.cell(row=row_num, column=2, value=pwd)
            ws.cell(row=row_num, column=3, value=ck)
            action = "UPDATED"
        else:
            new_row = ws.max_row + 1
            ws.cell(row=new_row, column=1, value=uid)
            ws.cell(row=new_row, column=2, value=pwd)
            ws.cell(row=new_row, column=3, value=ck)
            action = "ADDED"
        
        wb.save(path)
        total = ws.max_row
        
        count_1000 = get_fb_count(user_id, "1000X")
        count_6155 = get_fb_count(user_id, "6155X")
        fb_menu = ReplyKeyboardMarkup([
            [KeyboardButton(f"1000X({count_1000})", style="success", icon_custom_emoji_id=CUSTOM_EMOJI["📂"]), 
             KeyboardButton(f"6155X({count_6155})", style="primary", icon_custom_emoji_id=CUSTOM_EMOJI["📂"])],
            [KeyboardButton("Set Password", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🔒"])],
            [KeyboardButton("Download", style="success", icon_custom_emoji_id=CUSTOM_EMOJI["📥"])],
            [KeyboardButton("Reset Data", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🗑️"])],
            [KeyboardButton("BACK TO MAIN MENU", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🔙"])]
        ], resize_keyboard=True)
        
        await update.message.reply_text(
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["✅"]}">✅</tg-emoji> <b>DATA {action} SUCCESSFULLY!</b>\n\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["📂"]}">📂</tg-emoji> <b>File:</b> {opt}\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["📊"]}">📊</tg-emoji> <b>Total:</b> {total} entries\n\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["👑"]}">👑</tg-emoji> <b>FACEBOOK DATA MANAGER</b>\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["📊"]}">📊</tg-emoji> <b>Please select an option</b>',
            parse_mode='HTML', reply_markup=fb_menu
        )
        return FB_SELECTING_OPTION
    
    else:  # ❌ No
        count_1000 = get_fb_count(user_id, "1000X")
        count_6155 = get_fb_count(user_id, "6155X")
        fb_menu = ReplyKeyboardMarkup([
            [KeyboardButton(f"1000X({count_1000})", style="success", icon_custom_emoji_id=CUSTOM_EMOJI["📂"]), 
             KeyboardButton(f"6155X({count_6155})", style="primary", icon_custom_emoji_id=CUSTOM_EMOJI["📂"])],
            [KeyboardButton("Set Password", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🔒"])],
            [KeyboardButton("Download", style="success", icon_custom_emoji_id=CUSTOM_EMOJI["📥"])],
            [KeyboardButton("Reset Data", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🗑️"])],
            [KeyboardButton("BACK TO MAIN MENU", style="danger", icon_custom_emoji_id=CUSTOM_EMOJI["🔙"])]
        ], resize_keyboard=True)
        
        await update.message.reply_text(
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["❌"]}">❌</tg-emoji> <b>OPERATION CANCELLED</b>\n\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["👑"]}">👑</tg-emoji> <b>FACEBOOK DATA MANAGER</b>\n'
            f'<tg-emoji emoji-id="{CUSTOM_EMOJI["📊"]}">📊</tg-emoji> <b>Please select an option</b>',
            parse_mode='HTML', reply_markup=fb_menu
        )
        return FB_SELECTING_OPTION  # Facebook মেনুতেই থাকবে


# প্রথমে ফাইলের শুরুতে (অন্যান্য async ফাংশনের জায়গায়) এই ফাংশনটি যোগ করুন:
async def fb_start_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Facebook কনভারসেশন চলাকালীন /start দিলে মেইন মেনু দেখাবে"""
    context.user_data.clear()
    await update.message.reply_text(
        f'<tg-emoji emoji-id="{CUSTOM_EMOJI["🏠"]}">🏠</tg-emoji> <b>Back to Main Menu</b>',
        parse_mode='HTML',
        reply_markup=MAIN_MENU
    )
    return ConversationHandler.END

# তারপর নিচের ফাংশনটি পরিবর্তন করুন:
def facebook_conversation_handler():
    return ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^FACEBOOK COOKIES$"), fb_start)],
        states={
            FB_SELECTING_OPTION: [MessageHandler(filters.TEXT & ~filters.COMMAND, fb_select_option)],
            FB_SETTING_PASSWORD: [MessageHandler(filters.TEXT & ~filters.COMMAND, fb_set_password)],
            FB_WAITING_UID: [MessageHandler(filters.TEXT & ~filters.COMMAND, fb_waiting_uid)],
            FB_WAITING_PASSWORD: [MessageHandler(filters.TEXT & ~filters.COMMAND, fb_waiting_password)],
            FB_WAITING_PASSWORD_INPUT: [MessageHandler(filters.TEXT & ~filters.COMMAND, fb_waiting_password_input)],
            FB_WAITING_COOKIES: [MessageHandler(filters.TEXT & ~filters.COMMAND, fb_waiting_cookies)],
            FB_CONFIRM_SAVE: [MessageHandler(filters.TEXT & ~filters.COMMAND, fb_confirm_save)],
        },
        fallbacks=[
            CommandHandler("cancel", cancel),
            CommandHandler("start", fb_start_command),  # এই লাইনটি যোগ করুন
        ],
    )
