import re
import os
import json
import random
import shutil
import openpyxl
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import Application, CommandHandler, MessageHandler, filters, CallbackQueryHandler, ConversationHandler

# ==================== কনভার্সেশন স্টেটস ====================
WAITING_FOR_RENAME_FILE = 1
WAITING_FOR_NEW_FILENAME = 2

FB_SELECTING_OPTION = 10
FB_SETTING_PASSWORD = 11
FB_WAITING_UID = 12
FB_WAITING_PASSWORD = 13
FB_WAITING_PASSWORD_INPUT = 14
FB_WAITING_COOKIES = 15
FB_CONFIRM_SAVE = 16

# ==================== পাথ ডিটেক্ট ====================
def get_base_dir():
    if 'com.termux' in os.environ.get('PREFIX', '') or os.path.exists('/sdcard'):
        return "/sdcard/File_conv"
    else:
        return "/tmp/bot_files"

BASE_DIR = get_base_dir()
os.makedirs(BASE_DIR, exist_ok=True)

# ==================== কান্ট্রি ডাটা ====================
def load_country_data():
    try:
        with open('countries.json', 'r', encoding='utf-8') as f:
            return json.load(f)
    except:
        return {}

COUNTRY_DATA = load_country_data()

# ==================== মেইন মেনু ====================
MAIN_MENU = ReplyKeyboardMarkup([
    [KeyboardButton("📱 FACEBOOK COOKIES", style="primary")],
    [KeyboardButton("📁 FILE CHANGE", style="danger")],
    [KeyboardButton("🌍 NUMBER CONVERTER", style="success")],
    [KeyboardButton("🔄 RESET", style="danger"), KeyboardButton("ℹ️ HELP", style="success")]
], resize_keyboard=True)

# BACK ফাংশনালিটি - Facebook মেনুতে ফিরে যাওয়ার জন্য
BACK_FB_MENU = ReplyKeyboardMarkup([
    [KeyboardButton("🔙 BACK", style="danger")]
], resize_keyboard=True)

# ==================== নাম্বার কনভার্টার ফাংশন ====================
def clean_numbers_from_text(text):
    raw_numbers = re.findall(r'\b\d{9,15}\b', text)
    clean = []
    
    for num in raw_numbers:
        if len(num) == 10:
            if num.startswith('01'):
                num = '+880' + num[1:]
            else:
                num = '+91' + num
        elif len(num) == 11:
            code = num[:2]
            if code in COUNTRY_DATA:
                num = '+' + num
            else:
                code = num[:3]
                if code in COUNTRY_DATA:
                    num = '+' + num
                else:
                    num = '+' + num
        elif len(num) == 12:
            code = num[:3]
            if code in COUNTRY_DATA:
                num = '+' + num
            else:
                code = num[:2]
                if code in COUNTRY_DATA:
                    num = '+' + num
                else:
                    num = '+' + num
        elif len(num) == 13:
            code = num[:3]
            if code in COUNTRY_DATA:
                num = '+' + num
            else:
                num = '+' + num
        elif 11 <= len(num) <= 15:
            for code_len in [3, 2]:
                if len(num) > code_len:
                    code = num[:code_len]
                    if code in COUNTRY_DATA:
                        num = '+' + num
                        break
        
        if num.startswith('+') and num not in clean:
            clean.append(num)
    
    return clean

def read_txt_content(file_path):
    content = ""
    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
        for line in f:
            line = line.strip()
            if line and re.search(r'\d', line):
                content += line + " "
    return content

def read_xlsx_content(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    content = ""
    for row in sheet.iter_rows(values_only=True):
        for cell in row:
            if cell is not None:
                cell_str = str(cell).strip()
                if re.search(r'\d{9,}', cell_str):
                    content += cell_str + " "
    return content

def read_csv_content(file_path):
    content = ""
    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
        for line in f:
            parts = line.replace(',', ' ').split()
            for part in parts:
                if re.search(r'\d{9,}', part):
                    content += part + " "
    return content

def read_file_content(file_path, file_ext):
    if file_ext == '.xlsx':
        return read_xlsx_content(file_path)
    elif file_ext == '.csv':
        return read_csv_content(file_path)
    else:
        return read_txt_content(file_path)

def get_country_info(number):
    for code, info in COUNTRY_DATA.items():
        if number.startswith('+' + code):
            return {
                'code': code,
                'short_code': info.get('code', '??'),
                'flag': info['flag'],
                'name_en': info['name_en']
            }
    return {'code': '??', 'short_code': '??', 'flag': '🌍', 'name_en': 'Unknown'}

# ==================== ফেসবুক কুকিজ ফাংশন ====================
def get_fb_file_path(user_id, option):
    if option == "1000X":
        return os.path.join(BASE_DIR, f"{user_id}_1000X.xlsx")
    else:
        return os.path.join(BASE_DIR, f"{user_id}_6155X.xlsx")

def get_fb_count(user_id, option):
    path = get_fb_file_path(user_id, option)
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        return wb.active.max_row
    return 0

def validate_fb_uid(uid, option):
    if option == "1000X":
        return uid.startswith("1000")
    else:
        return uid.startswith(("6155", "6156", "6157", "6158"))

# ==================== /cancel কমান্ড ====================
async def cancel(update: Update, context):
    context.user_data.clear()
    await update.message.reply_text(
        "✨ Cancelled successfully!\n"
        "No worries — returning to Main Menu 🏠",
        reply_markup=MAIN_MENU
    )
    return ConversationHandler.END

# ==================== স্টার্ট ====================
async def start(update: Update, context):
    await update.message.reply_text(
        "🔥 *BEST NUMBER TOOL ON TELEGRAM* 🔥\n\n"
        "✅ Clean 10,000+ numbers in seconds\n"
        "✅ 100% accurate country detection\n"
        "✅ No data stored (except FB manager)\n\n"
        "🚀 *Just send a file & relax!*\n\n"
        f"📞 *{len(COUNTRY_DATA)} countries* | 📁 3 formats",
        parse_mode="Markdown",
        reply_markup=MAIN_MENU
    )

# ==================== মেইন হ্যান্ডলার ====================
async def handle_menu(update: Update, context):
    text = update.message.text
    
    # 🔙 BACK বাটন হ্যান্ডলার - NUMBER CONVERTER বা FILE CHANGE থেকে BACK করলে
    if text == "🔙 BACK":
        context.user_data.clear()
        await update.message.reply_text(
            "🔙 *Returning to Main Menu*",
            parse_mode="Markdown",
            reply_markup=MAIN_MENU
        )
        return
    
    if text == "🔙 BACK TO MAIN MENU":
        context.user_data.clear()
        await update.message.reply_text(
            "🏠 *Main Menu*\n"
            "━━━━━━━━━━\n"
            "✅ Session ended\n"
            "👇 Choose an option:",
            parse_mode="Markdown",
            reply_markup=MAIN_MENU
        )
        return ConversationHandler.END
    
    # নাম্বার কনভার্টার
    if text == "🌍 NUMBER CONVERTER":
        context.user_data.clear()
        context.user_data['mode'] = 'number_converter'
        await update.message.reply_text(
            "📎 *FILE UPLOAD REQUIRED*\n\n"
            "┌─────────────────────┐\n"
            "│ ✅ Supported formats │\n"
            "├─────────────────────┤\n"
            "│ • .txt              │\n"
            "│ • .csv              │\n"
            "│ • .xlsx             │\n"
            "└─────────────────────┘\n\n"
            "📤 *Send your file now*\n\n"
            "🔙 Press BACK to cancel",
            parse_mode="Markdown",
            reply_markup=BACK_FB_MENU
        )
        return
    
    # ফাইল রিনেমার
    if text == "📁 FILE CHANGE":
        context.user_data.clear()
        context.user_data['mode'] = 'file_renamer'
        await update.message.reply_text(
            "✏️ *FILE RENAMER TOOL*\n\n"
            "📤 *Step 1:* Send file\n"
            "✏️ *Step 2:* Type new name\n"
            "✅ *Step 3:* Receive renamed file\n\n"
            "📁 Supported: .txt | .csv | .xlsx\n\n"
            "👉 *Send your file now*\n\n"
            "🔙 Press BACK to cancel",
            parse_mode="Markdown",
            reply_markup=BACK_FB_MENU
        )
        return WAITING_FOR_RENAME_FILE
    
    # ফেসবুক কুকিজ
    if text == "📱 FACEBOOK COOKIES":
        user_id = update.effective_user.id
        context.user_data['fb_user_id'] = user_id
        count_1000 = get_fb_count(user_id, "1000X")
        count_6155 = get_fb_count(user_id, "6155X")
        
        fb_menu = ReplyKeyboardMarkup([
            [KeyboardButton(f"📁 1000X ({count_1000})", style="success"), KeyboardButton(f"📁 6155X ({count_6155})", style="primary")],
            [KeyboardButton("🔒 Set Password", style="danger"), KeyboardButton("📥 Download", style="success")],
            [KeyboardButton("⚙️ Settings", style="primary")],
            [KeyboardButton("🔙 BACK TO MAIN MENU", style="danger")]
        ], resize_keyboard=True)
        
        await update.message.reply_text(
            "✨ *━━━━━━━━━━━━━━━━━━━━━━━━━* ✨\n"
            "👑 *VIP FACEBOOK DATA MANAGER* 👑\n"
            "✨ *━━━━━━━━━━━━━━━━━━━━━━━━━* ✨\n\n"
            "💎 *Welcome to Team X Bot* 💎\n"
            "🎯 *System Ready*\n\n"
            "📊 *Please select an option below:*",
            parse_mode="Markdown",
            reply_markup=fb_menu
        )
        return FB_SELECTING_OPTION
    
    if text == "🔄 RESET":
        context.user_data.clear()
        await update.message.reply_text(
            "┌─────────────────┐\n"
            "│ 🗑 DATA CLEARED │\n"
            "├─────────────────┤\n"
            "│ ✅ Ready        │\n"
            "│ 🏠 Main Menu    │\n"
            "└─────────────────┘",
            parse_mode="Markdown",
            reply_markup=MAIN_MENU
        )
        return
    
    if text == "ℹ️ HELP":
        await update.message.reply_text(
            "📖 *COMMAND REFERENCE*\n\n"
            "┌─────────────────────────┐\n"
            "│ 🌍 NUMBER CONVERTER     │\n"
            "│ → Extract + country flag│\n"
            "├─────────────────────────┤\n"
            "│ 📁 FILE CHANGE          │\n"
            "│ → Rename files          │\n"
            "├─────────────────────────┤\n"
            "│ 📱 FACEBOOK COOKIES     │\n"
            "│ → Store UID/Pass/Cookies│\n"
            "├─────────────────────────┤\n"
            "│ 🔄 RESET                │\n"
            "│ → Clear temp data       │\n"
            "└─────────────────────────┘\n\n"
            f"🌍 *{len(COUNTRY_DATA)} countries supported*",
            parse_mode="Markdown",
            reply_markup=MAIN_MENU
        )
        return

# ==================== নাম্বার কনভার্টার ফাইল হ্যান্ডলার ====================
async def handle_number_file(update: Update, context):
    # 🔙 BACK বাটন হ্যান্ডলার
    if update.message.text == "🔙 BACK":
        context.user_data.clear()
        await update.message.reply_text(
            "┌─────────────────────┐\n"
            "│  🔙 BACK TO MENU    │\n"
            "└─────────────────────┘",
            parse_mode="Markdown",
            reply_markup=MAIN_MENU
        )
        return
    
    if context.user_data.get('mode') != 'number_converter':
        return
    
    if 'random_mode' not in context.user_data:
        context.user_data['random_mode'] = False
    
    processing_msg = await update.message.reply_text(
        "┌─────────────────────┐\n"
        "│  ⏳ PROCESSING      │\n"
        "│  📁 Please wait...  │\n"
        "└─────────────────────┘",
        parse_mode="Markdown"
    )
    
    file = await update.message.document.get_file()
    file_name = update.message.document.file_name
    
    if file_name.endswith('.xlsx'):
        file_ext = '.xlsx'
    elif file_name.endswith('.csv'):
        file_ext = '.csv'
    else:
        file_ext = '.txt'
    
    file_path = os.path.join(BASE_DIR, f"{file.file_id}{file_ext}")
    await file.download_to_drive(file_path)
    
    try:
        content = read_file_content(file_path, file_ext)
    except Exception as e:
        await processing_msg.edit_text(
            f"┌─────────────────────┐\n"
            f"│  ❌ ERROR           │\n"
            f"│  {str(e)[:20]}│\n"
            f"└─────────────────────┘",
            parse_mode="Markdown"
        )
        os.remove(file_path)
        return
    
    numbers = clean_numbers_from_text(content)
    
    if not numbers:
        await processing_msg.edit_text(
            "┌─────────────────────────────────┐\n"
            "│  ❌ NO VALID NUMBERS FOUND!     │\n"
            "├─────────────────────────────────┤\n"
            "│  📞 CORRECT FORMAT:              │\n"
            "│  • 965XXXXXXXXX (Kuwait)        │\n"
            "│  • 998XXXXXXXXX (Uzbekistan)    │\n"
            "│  • 93XXXXXXXXX (Afghanistan)    │\n"
            "│  • 01XXXXXXXXX (Bangladesh)     │\n"
            "│  • 98XXXXXXXX (India)           │\n"
            "└─────────────────────────────────┘",
            parse_mode="Markdown"
        )
        os.remove(file_path)
        return
    
    country_info = get_country_info(numbers[0])
    context.user_data['country_info'] = country_info
    main_country = f"{country_info['flag']} {country_info['name_en']} ({country_info['short_code']}) +{country_info['code']}"
    
    context.user_data['numbers'] = numbers
    random_mode_status = "🟢 ON" if context.user_data['random_mode'] else "🔴 OFF"
    
    await processing_msg.delete()
    
    inline_keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton(f"🎲 Random Mode: {random_mode_status}", callback_data='toggle_random')],
        [InlineKeyboardButton(f"✅ Convert ({len(numbers)})", callback_data='num_convert'),
         InlineKeyboardButton(f"➕ Convert with +", callback_data='num_convert_plus')],
        [InlineKeyboardButton(f"🔄 Reset", callback_data='num_reset')]
    ])
    
    await update.message.reply_text(
        f"╔══════════════════════════════╗\n"
        f"║     📊 PROCESSING COMPLETE    ║\n"
        f"╠══════════════════════════════╣\n"
        f"║  🌍 Country: {main_country[:25]}\n"
        f"║  🔢 Valid Numbers: {len(numbers)}\n"
        f"║  📝 Sample: `{numbers[0]}`\n"
        f"║  🎲 Random Mode: {random_mode_status}\n"
        f"║  📌 Output: +{country_info['code']}XXXXXXXXX\n"
        f"╚══════════════════════════════╝",
        reply_markup=inline_keyboard,
        parse_mode="Markdown"
    )
    
    os.remove(file_path)

async def toggle_random_callback(update: Update, context):
    query = update.callback_query
    await query.answer()
    
    current_mode = context.user_data.get('random_mode', False)
    context.user_data['random_mode'] = not current_mode
    
    numbers = context.user_data.get('numbers', [])
    country_info = context.user_data.get('country_info', {'code': '??', 'flag': '🌍', 'name_en': 'Unknown', 'short_code': '??'})
    random_mode_status = "🟢 ON" if context.user_data['random_mode'] else "🔴 OFF"
    
    inline_keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton(f"🎲 Random Mode: {random_mode_status}", callback_data='toggle_random')],
        [InlineKeyboardButton(f"✅ Convert ({len(numbers)})", callback_data='num_convert'),
         InlineKeyboardButton(f"➕ Convert with +", callback_data='num_convert_plus')],
        [InlineKeyboardButton(f"🔄 Reset", callback_data='num_reset')]
    ])
    
    main_country = f"{country_info['flag']} {country_info['name_en']} ({country_info['short_code']}) +{country_info['code']}"
    
    await query.edit_message_text(
        f"╔══════════════════════════════╗\n"
        f"║     📊 PROCESSING COMPLETE    ║\n"
        f"╠══════════════════════════════╣\n"
        f"║  🌍 Country: {main_country[:25]}\n"
        f"║  🔢 Valid Numbers: {len(numbers)}\n"
        f"║  📝 Sample: `{numbers[0]}`\n"
        f"║  🎲 Random Mode: {random_mode_status}\n"
        f"║  📌 Output: +{country_info['code']}XXXXXXXXX\n"
        f"╚══════════════════════════════╝",
        reply_markup=inline_keyboard,
        parse_mode="Markdown"
    )

async def num_convert_callback(update: Update, context):
    query = update.callback_query
    await query.answer()
    
    numbers = context.user_data.get('numbers', [])
    random_mode = context.user_data.get('random_mode', False)
    
    if not numbers:
        await query.message.reply_text(
            "┌─────────────────────┐\n"
            "│  ❌ NO NUMBERS      │\n"
            "│  📁 File not found  │\n"
            "└─────────────────────┘",
            parse_mode="Markdown"
        )
        return
    
    country_info = context.user_data.get('country_info', {'name_en': 'Unknown', 'code': '??'})
    
    if random_mode:
        shuffled_numbers = numbers.copy()
        random.shuffle(shuffled_numbers)
        output = "\n".join(shuffled_numbers)
        mode_text = "🎲 Random Order"
    else:
        output = "\n".join(numbers)
        mode_text = "📋 Serial Order"
    
    filename = f"{country_info['name_en']}_Numbers.txt"
    filepath = os.path.join(BASE_DIR, filename)
    
    with open(filepath, "w") as f:
        f.write(output)
    
    with open(filepath, "rb") as f:
        await query.message.reply_document(
            document=f,
            filename=filename,
            caption=f"╔══════════════════════════════╗\n"
                    f"║     ✅ CONVERSION COMPLETE    ║\n"
                    f"╠══════════════════════════════╣\n"
                    f"║  🔢 Numbers: {len(numbers)}\n"
                    f"║  🎲 Mode: {mode_text}\n"
                    f"║  📌 Format: +{country_info['code']}XXXXXXXXX\n"
                    f"╚══════════════════════════════╝"
        )
    
    os.remove(filepath)

async def num_convert_plus_callback(update: Update, context):
    query = update.callback_query
    await query.answer()
    
    numbers = context.user_data.get('numbers', [])
    random_mode = context.user_data.get('random_mode', False)
    
    if not numbers:
        await query.message.reply_text(
            "┌─────────────────────┐\n"
            "│  ❌ NO NUMBERS      │\n"
            "│  📁 File not found  │\n"
            "└─────────────────────┘",
            parse_mode="Markdown"
        )
        return
    
    country_info = context.user_data.get('country_info', {'name_en': 'Unknown', 'code': '??'})
    
    if random_mode:
        shuffled_numbers = numbers.copy()
        random.shuffle(shuffled_numbers)
        output = "\n".join([f"+{num}" for num in shuffled_numbers])  # + যোগ করা হলো
        mode_text = "🎲 Random Order"
    else:
        output = "\n".join([f"+{num}" for num in numbers])  # + যোগ করা হলো
        mode_text = "📋 Serial Order"
    
    filename = f"{country_info['name_en']}_Numbers_With_Plus.txt"
    filepath = os.path.join(BASE_DIR, filename)
    
    with open(filepath, "w") as f:
        f.write(output)
    
    with open(filepath, "rb") as f:
        await query.message.reply_document(
            document=f,
            filename=filename,
            caption=f"╔══════════════════════════════╗\n"
                    f"║  ✅ CONVERSION (+ FORMAT)    ║\n"
                    f"╠══════════════════════════════╣\n"
                    f"║  🔢 Numbers: {len(numbers)}\n"
                    f"║  🎲 Mode: {mode_text}\n"
                    f"║  📌 Format: +{country_info['code']}XXXXXXXXX\n"
                    f"╚══════════════════════════════╝"
        )
    
    os.remove(filepath)

async def num_reset_callback(update: Update, context):
    query = update.callback_query
    await query.answer()
    
    context.user_data.pop('numbers', None)
    context.user_data.pop('random_mode', None)
    context.user_data.pop('country_info', None)
    
    await query.message.reply_text(
        "┌─────────────────────┐\n"
        "│  🔄 RESET DONE      │\n"
        "├─────────────────────┤\n"
        "│  🗑 Data cleared    │\n"
        "│  📁 Ready for new   │\n"
        "│  📤 Send file now   │\n"
        "└─────────────────────┘",
        parse_mode="Markdown",
        reply_markup=MAIN_MENU
    )

# ==================== ফাইল রিনেমার হ্যান্ডলার ====================
async def handle_rename_file(update: Update, context):
    text = update.message.text
    
    if text == "🔙 BACK":
        # টেম্প ফাইল থাকলে ডিলিট করুন
        if 'original_file_path' in context.user_data:
            path = context.user_data['original_file_path']
            if os.path.exists(path):
                os.remove(path)
        context.user_data.clear()
        await update.message.reply_text(
            "┌─────────────────────┐\n"
            "│  🔙 BACK TO MENU    │\n"
            "└─────────────────────┘",
            parse_mode="Markdown",
            reply_markup=MAIN_MENU
        )
        return ConversationHandler.END
    
    if not update.message.document:
        await update.message.reply_text(
            "┌─────────────────────────────┐\n"
            "│  ❌ NO FILE DETECTED        │\n"
            "├─────────────────────────────┤\n"
            "│  📤 Send a valid file       │\n"
            "│  📁 .txt | .csv | .xlsx     │\n"
            "├─────────────────────────────┤\n"
            "│  🔙 Or press BACK to cancel │\n"
            "└─────────────────────────────┘",
            parse_mode="Markdown",
            reply_markup=BACK_FB_MENU
        )
        return WAITING_FOR_RENAME_FILE
    
    file = await update.message.document.get_file()
    file_name = update.message.document.file_name
    
    if file_name.endswith('.xlsx'):
        file_ext = '.xlsx'
    elif file_name.endswith('.csv'):
        file_ext = '.csv'
    elif file_name.endswith('.txt'):
        file_ext = '.txt'
    else:
        await update.message.reply_text(
            "┌─────────────────────────────┐\n"
            "│  ❌ UNSUPPORTED FORMAT      │\n"
            "├─────────────────────────────┤\n"
            "│  ✅ Supported formats:      │\n"
            "│  • .txt                     │\n"
            "│  • .csv                     │\n"
            "│  • .xlsx                    │\n"
            "└─────────────────────────────┘",
            parse_mode="Markdown",
            reply_markup=BACK_FB_MENU
        )
        return WAITING_FOR_RENAME_FILE
    
    original_path = os.path.join(BASE_DIR, f"original_{file.file_id}{file_ext}")
    await file.download_to_drive(original_path)
    
    context.user_data['original_file_path'] = original_path
    context.user_data['original_file_ext'] = file_ext
    context.user_data['original_file_name'] = file_name
    
    await update.message.reply_text(
        f"┌─────────────────────────────┐\n"
        f"│  📄 FILE RECEIVED           │\n"
        f"├─────────────────────────────┤\n"
        f"│  📛 Name: `{file_name[:20]}`│\n"
        f"├─────────────────────────────┤\n"
        f"│  ✏️ Send new file name      │\n"
        f"│  (without extension)        │\n"
        f"├─────────────────────────────┤\n"
        f"│  📝 Example:                │\n"
        f"│  `my_renamed_file`          │\n"
        f"├─────────────────────────────┤\n"
        f"│  🔙 Press BACK to cancel    │\n"
        f"└─────────────────────────────┘",
        parse_mode="Markdown",
        reply_markup=BACK_FB_MENU
    )
    return WAITING_FOR_NEW_FILENAME

async def handle_new_filename(update: Update, context):
    text = update.message.text.strip()
    
    if text == "🔙 BACK":
        if 'original_file_path' in context.user_data:
            path = context.user_data['original_file_path']
            if os.path.exists(path):
                os.remove(path)
        await update.message.reply_text(
            "┌─────────────────────┐\n"
            "│  🔙 BACK TO MENU    │\n"
            "└─────────────────────┘",
            parse_mode="Markdown",
            reply_markup=MAIN_MENU
        )
        return ConversationHandler.END
    
    safe_name = re.sub(r'[<>:"/\\|?*]', '_', text).replace(' ', '_')
    
    if not safe_name:
        await update.message.reply_text(
            "┌─────────────────────┐\n"
            "│  ❌ INVALID NAME    │\n"
            "├─────────────────────┤\n"
            "│  ✏️ Try another     │\n"
            "│  📝 Use letters &   │\n"
            "│     numbers only    │\n"
            "└─────────────────────┘",
            parse_mode="Markdown"
        )
        return WAITING_FOR_NEW_FILENAME
    
    original_path = context.user_data.get('original_file_path')
    file_ext = context.user_data.get('original_file_ext', '.txt')
    
    if not original_path or not os.path.exists(original_path):
        await update.message.reply_text(
            "┌─────────────────────┐\n"
            "│  ❌ FILE NOT FOUND  │\n"
            "├─────────────────────┤\n"
            "│  🔄 Please restart  │\n"
            "│     the process     │\n"
            "└─────────────────────┘",
            parse_mode="Markdown",
            reply_markup=MAIN_MENU
        )
        return ConversationHandler.END
    
    new_filename = f"{safe_name}{file_ext}"
    new_path = os.path.join(BASE_DIR, new_filename)
    
    shutil.copy2(original_path, new_path)
    
    with open(new_path, "rb") as f:
        await update.message.reply_document(
            document=f,
            filename=new_filename,
            caption=f"╔══════════════════════════════════╗\n"
                    f"║     ✅ FILE RENAMED SUCCESSFULLY   ║\n"
                    f"╠══════════════════════════════════╣\n"
                    f"║  📄 Original: `{context.user_data.get('original_file_name', 'Unknown')[:20]}`\n"
                    f"║  📄 New: `{new_filename[:20]}`\n"
                    f"╚══════════════════════════════════╝",
            parse_mode="Markdown"
        )
    
    os.remove(original_path)
    os.remove(new_path)
    
    context.user_data.clear()
    await update.message.reply_text(
        "┌─────────────────────┐\n"
        "│  🏠 BACK TO MENU    │\n"
        "└─────────────────────┘",
        parse_mode="Markdown",
        reply_markup=MAIN_MENU
    )
    return ConversationHandler.END

# ==================== ফেসবুক কুকিজ হ্যান্ডলার ====================
async def fb_select_option(update: Update, context):
    text = update.message.text
    user_id = context.user_data.get('fb_user_id')
    
    # 🔙 BACK বাটন হ্যান্ডলার - Facebook মেনুতে ফিরে যাবে
    if text == "🔙 BACK":
        count_1000 = get_fb_count(user_id, "1000X")
        count_6155 = get_fb_count(user_id, "6155X")
        
        fb_menu = ReplyKeyboardMarkup([
            [KeyboardButton(f"📁 1000X ({count_1000})", style="success"), KeyboardButton(f"📁 6155X ({count_6155})", style="primary")],
            [KeyboardButton("🔒 Set Password", style="danger"), KeyboardButton("📥 Download", style="success")],
            [KeyboardButton("⚙️ Settings", style="primary")],
            [KeyboardButton("🔙 BACK TO MAIN MENU", style="danger")]
        ], resize_keyboard=True)
        
        await update.message.reply_text(
            "┌─────────────────────────────┐\n"
            "│  👑 FACEBOOK DATA MANAGER   │\n"
            "├─────────────────────────────┤\n"
            "│  📊 Please select an option │\n"
            "└─────────────────────────────┘",
            parse_mode="Markdown",
            reply_markup=fb_menu
        )
        return FB_SELECTING_OPTION
    
    if text == "🔙 BACK TO MAIN MENU":
        await update.message.reply_text(
            "┌─────────────────────┐\n"
            "│  🏠 BACK TO MENU    │\n"
            "└─────────────────────┘",
            parse_mode="Markdown",
            reply_markup=MAIN_MENU
        )
        return ConversationHandler.END
    
    if "1000X" in text:
        context.user_data['fb_current_option'] = "1000X"
        path = get_fb_file_path(user_id, "1000X")
        if not os.path.exists(path):
            wb = openpyxl.Workbook()
            wb.save(path)
        await update.message.reply_text(
            "┌─────────────────────────────┐\n"
            "│  📱 ENTER FB ACCOUNT UID    │\n"
            "├─────────────────────────────┤\n"
            "│  📝 Format:                 │\n"
            "│  Must start with 1000       │\n"
            "│  Example: 1000123456789     │\n"
            "├─────────────────────────────┤\n"
            "│  🔙 BACK to cancel          │\n"
            "└─────────────────────────────┘",
            parse_mode="Markdown",
            reply_markup=BACK_FB_MENU
        )
        return FB_WAITING_UID
    
    elif "6155X" in text:
        context.user_data['fb_current_option'] = "6155X"
        path = get_fb_file_path(user_id, "6155X")
        if not os.path.exists(path):
            wb = openpyxl.Workbook()
            wb.save(path)
        await update.message.reply_text(
            "┌─────────────────────────────┐\n"
            "│  📱 ENTER FB ACCOUNT UID    │\n"
            "├─────────────────────────────┤\n"
            "│  📝 Format:                 │\n"
            "│  Must start with:           │\n"
            "│  6155, 6156, 6157, 6158    │\n"
            "│  Example: 6155123456789     │\n"
            "├─────────────────────────────┤\n"
            "│  🔙 BACK to cancel          │\n"
            "└─────────────────────────────┘",
            parse_mode="Markdown",
            reply_markup=BACK_FB_MENU
        )
        return FB_WAITING_UID
    
    elif text == "🔒 Set Password":
        await update.message.reply_text(
            "┌─────────────────────────────┐\n"
            "│  🔐 SET GLOBAL PASSWORD     │\n"
            "├─────────────────────────────┤\n"
            "│  Enter your password        │\n"
            "│  Type 'skip' to continue    │\n"
            "│  without password           │\n"
            "├─────────────────────────────┤\n"
            "│  🔙 BACK to cancel          │\n"
            "└─────────────────────────────┘",
            parse_mode="Markdown",
            reply_markup=BACK_FB_MENU
        )
        return FB_SETTING_PASSWORD
    
    elif text == "⚙️ Settings":
        settings_menu = ReplyKeyboardMarkup([
            [KeyboardButton("🗑️ Reset Data", style="primary")],
            [KeyboardButton("💾 Save Data", style="success")],
            [KeyboardButton("🔙 BACK", style="danger")]
        ], resize_keyboard=True)
        await update.message.reply_text(
            "┌─────────────────────────────┐\n"
            "│  ⚙️ SETTINGS PANEL          │\n"
            "├─────────────────────────────┤\n"
            "│  Configure your preferences │\n"
            "└─────────────────────────────┘",
            parse_mode="Markdown",
            reply_markup=settings_menu
        )
        return FB_SELECTING_OPTION
    
    elif text == "🗑️ Reset Data":
        for opt in ["1000X", "6155X"]:
            path = get_fb_file_path(user_id, opt)
            if os.path.exists(path):
                os.remove(path)
        await update.message.reply_text(
            "┌─────────────────────────────┐\n"
            "│  ✅ DATA CLEARED            │\n"
            "│  All files have been deleted│\n"
            "└─────────────────────────────┘",
            parse_mode="Markdown",
            reply_markup=MAIN_MENU
        )
        return ConversationHandler.END
    
    elif text == "💾 Save Data":
        await update.message.reply_text(
            "┌─────────────────────────────┐\n"
            "│  💾 SAVED                   │\n"
            "│  All changes saved          │\n"
            "└─────────────────────────────┘",
            parse_mode="Markdown",
            reply_markup=MAIN_MENU
        )
        return ConversationHandler.END
    
    elif text == "📥 Download":
        download_menu = ReplyKeyboardMarkup([
            [KeyboardButton("📁 1000X File", style="success"), KeyboardButton("📁 6155X File", style="primary")],
            [KeyboardButton("📁 Both Files", style="success")],
            [KeyboardButton("🔙 BACK", style="danger")]
        ], resize_keyboard=True)
        await update.message.reply_text(
            "┌─────────────────────────────┐\n"
            "│  📥 DOWNLOAD CENTER         │\n"
            "├─────────────────────────────┤\n"
            "│  Select file to download:   │\n"
            "└─────────────────────────────┘",
            parse_mode="Markdown",
            reply_markup=download_menu
        )
        return FB_SELECTING_OPTION
    
    elif text == "📁 1000X File":
        return await fb_download_file(update, context, "1000X")
    elif text == "📁 6155X File":
        return await fb_download_file(update, context, "6155X")
    elif text == "📁 Both Files":
        return await fb_download_both(update, context)
    
    else:
        await update.message.reply_text(
            "┌─────────────────────────────┐\n"
            "│  ❌ INVALID OPTION          │\n"
            "├─────────────────────────────┤\n"
            "│  Please use the buttons     │\n"
            "└─────────────────────────────┘",
            parse_mode="Markdown"
        )
        return FB_SELECTING_OPTION

async def fb_download_file(update: Update, context, option):
    user_id = context.user_data.get('fb_user_id')
    path = get_fb_file_path(user_id, option)
    
    if not os.path.exists(path):
        await update.message.reply_text(
            "┌─────────────────────────────┐\n"
            f"│  ❌ NO DATA FOUND           │\n"
            "├─────────────────────────────┤\n"
            f"│  📁 {option} has no data    │\n"
            "├─────────────────────────────┤\n"
            "│  📤 Add data first          │\n"
            "└─────────────────────────────┘",
            parse_mode="Markdown"
        )
        return FB_SELECTING_OPTION
    
    wb = openpyxl.load_workbook(path)
    total = wb.active.max_row
    
    with open(path, "rb") as f:
        await update.message.reply_document(
            document=f,
            filename=f"{option}.xlsx",
            caption=f"╔══════════════════════════════╗\n"
                    f"║  ✅ FILE READY FOR DOWNLOAD  ║\n"
                    f"╠══════════════════════════════╣\n"
                    f"║  📁 File: {option:<16}║\n"
                    f"║  📊 Total: {total} entries             ║\n"
                    f"╚══════════════════════════════╝"
        )
    
    return FB_SELECTING_OPTION

async def fb_download_both(update: Update, context):
    user_id = context.user_data.get('fb_user_id')
    sent_count = 0
    
    for opt in ["1000X", "6155X"]:
        path = get_fb_file_path(user_id, opt)
        if os.path.exists(path):
            wb = openpyxl.load_workbook(path)
            total = wb.active.max_row
            with open(path, "rb") as f:
                await update.message.reply_document(
                    document=f, 
                    filename=f"{opt}.xlsx", 
                    caption=f"╔════════════════════════╗\n"
                            f"║  📁 {opt:<12}║\n"
                            f"║  📊 {total} entries        ║\n"
                            f"╚════════════════════════╝"
                )
            sent_count += 1
        else:
            await update.message.reply_text(
                f"┌─────────────────────────┐\n"
                f"│  ⚠️ NO DATA FOR {opt:<9}│\n"
                f"└─────────────────────────┘",
                parse_mode="Markdown"
            )
    
    if sent_count == 0:
        await update.message.reply_text(
            "┌─────────────────────────────┐\n"
            "│  ❌ NO FILES AVAILABLE      │\n"
            "├─────────────────────────────┤\n"
            "│  No data found in both      │\n"
            "│  1000X and 6155X folders    │\n"
            "└─────────────────────────────┘",
            parse_mode="Markdown"
        )
    
    return FB_SELECTING_OPTION

async def fb_set_password(update: Update, context):
    text = update.message.text
    user_id = context.user_data.get('fb_user_id')
    
    if text == "🔙 BACK":
        count_1000 = get_fb_count(user_id, "1000X")
        count_6155 = get_fb_count(user_id, "6155X")
        
        fb_menu = ReplyKeyboardMarkup([
            [KeyboardButton(f"📁 1000X ({count_1000})", style="success"), KeyboardButton(f"📁 6155X ({count_6155})", style="primary")],
            [KeyboardButton("🔒 Set Password", style="danger"), KeyboardButton("📥 Download", style="success")],
            [KeyboardButton("⚙️ Settings", style="primary")],
            [KeyboardButton("🔙 BACK TO MAIN MENU", style="danger")]
        ], resize_keyboard=True)
        
        await update.message.reply_text(
            "┌─────────────────────────────┐\n"
            "│  👑 FACEBOOK DATA MANAGER   │\n"
            "├─────────────────────────────┤\n"
            "│  📊 Please select an option │\n"
            "└─────────────────────────────┘",
            parse_mode="Markdown",
            reply_markup=fb_menu
        )
        return FB_SELECTING_OPTION
    
    if text.lower() == "skip":
        context.user_data['fb_global_password'] = None
        await update.message.reply_text(
            "┌─────────────────────────────┐\n"
            "│  ⏭️ PASSWORD SKIPPED        │\n"
            "├─────────────────────────────┤\n"
            "│  No password will be set    │\n"
            "└─────────────────────────────┘",
            parse_mode="Markdown"
        )
    else:
        context.user_data['fb_global_password'] = text
        updated_count = 0
        
        for opt in ["1000X", "6155X"]:
            path = get_fb_file_path(user_id, opt)
            if os.path.exists(path):
                wb = openpyxl.load_workbook(path)
                ws = wb.active
                for row in range(1, ws.max_row + 1):
                    ws.cell(row=row, column=2, value=text)
                wb.save(path)
                updated_count += 1
        
        await update.message.reply_text(
            f"╔══════════════════════════════╗\n"
            f"║  🔐 PASSWORD SET SUCCESSFULLY ║\n"
            f"╠══════════════════════════════╣\n"
            f"║  📁 Files updated: {updated_count}           ║\n"
            f"║  🔒 Password: `{text[:10]}...`    ║\n"
            f"╚══════════════════════════════╝",
            parse_mode="Markdown"
        )
    
    # পাসওয়ার্ড সেট করার পর Facebook মেনুতে ফিরে যান (মেইন মেনুতে না)
    count_1000 = get_fb_count(user_id, "1000X")
    count_6155 = get_fb_count(user_id, "6155X")
    
    fb_menu = ReplyKeyboardMarkup([
        [KeyboardButton(f"📁 1000X ({count_1000})", style="success"), KeyboardButton(f"📁 6155X ({count_6155})", style="primary")],
        [KeyboardButton("🔒 Set Password", style="danger"), KeyboardButton("📥 Download", style="success")],
        [KeyboardButton("⚙️ Settings", style="primary")],
        [KeyboardButton("🔙 BACK TO MAIN MENU", style="danger")]
    ], resize_keyboard=True)
    
    await update.message.reply_text(
        "┌─────────────────────────────┐\n"
        "│  👑 FACEBOOK DATA MANAGER   │\n"
        "├─────────────────────────────┤\n"
        "│  📊 Please select an option │\n"
        "└─────────────────────────────┘",
        parse_mode="Markdown",
        reply_markup=fb_menu
    )
    return FB_SELECTING_OPTION  # এটা ConversationHandler.END না, FB_SELECTING_OPTION হবে

async def fb_waiting_uid(update: Update, context):
    text = update.message.text.strip()
    current_option = context.user_data.get('fb_current_option')
    
    if text == "🔙 BACK":
        user_id = context.user_data.get('fb_user_id')
        count_1000 = get_fb_count(user_id, "1000X")
        count_6155 = get_fb_count(user_id, "6155X")
        
        fb_menu = ReplyKeyboardMarkup([
            [KeyboardButton(f"📁 1000X ({count_1000})", style="success"), KeyboardButton(f"📁 6155X ({count_6155})", style="primary")],
            [KeyboardButton("🔒 Set Password", style="danger"), KeyboardButton("📥 Download", style="success")],
            [KeyboardButton("⚙️ Settings", style="primary")],
            [KeyboardButton("🔙 BACK TO MAIN MENU", style="danger")]
        ], resize_keyboard=True)
        
        await update.message.reply_text(
            "┌─────────────────────────────┐\n"
            "│  👑 FACEBOOK DATA MANAGER   │\n"
            "├─────────────────────────────┤\n"
            "│  📊 Please select an option │\n"
            "└─────────────────────────────┘",
            parse_mode="Markdown",
            reply_markup=fb_menu
        )
        return FB_SELECTING_OPTION
    
    if not validate_fb_uid(text, current_option):
        await update.message.reply_text(
            f"┌─────────────────────────────────┐\n"
            f"│  ❌ INVALID UID FOR {current_option}  │\n"
            f"├─────────────────────────────────┤\n"
            f"│  📝 Please enter a valid UID:    │\n"
            f"│  🔙 BACK to cancel              │\n"
            f"└─────────────────────────────────┘",
            parse_mode="Markdown",
            reply_markup=BACK_FB_MENU
        )
        return FB_WAITING_UID
    
    context.user_data['fb_current_uid'] = text
    
    pwd_menu = ReplyKeyboardMarkup([
        [KeyboardButton("🔐 Use Global", style="success"), KeyboardButton("✏️ New Password", style="primary")],
        [KeyboardButton("⏭️ Skip", style="success")],
        [KeyboardButton("🔙 BACK", style="danger")]
    ], resize_keyboard=True)
    
    await update.message.reply_text(
        "┌─────────────────────────────┐\n"
        "│  🔐 PASSWORD OPTIONS        │\n"
        "├─────────────────────────────┤\n"
        "│  How would you like to set  │\n"
        "│  the password?              │\n"
        "└─────────────────────────────┘",
        parse_mode="Markdown",
        reply_markup=pwd_menu
    )
    return FB_WAITING_PASSWORD

async def fb_waiting_password(update: Update, context):
    text = update.message.text
    
    if text == "🔙 BACK":
        user_id = context.user_data.get('fb_user_id')
        count_1000 = get_fb_count(user_id, "1000X")
        count_6155 = get_fb_count(user_id, "6155X")
        
        fb_menu = ReplyKeyboardMarkup([
            [KeyboardButton(f"📁 1000X ({count_1000})", style="success"), KeyboardButton(f"📁 6155X ({count_6155})", style="primary")],
            [KeyboardButton("🔒 Set Password", style="danger"), KeyboardButton("📥 Download", style="success")],
            [KeyboardButton("⚙️ Settings", style="primary")],
            [KeyboardButton("🔙 BACK TO MAIN MENU", style="danger")]
        ], resize_keyboard=True)
        
        await update.message.reply_text(
            "┌─────────────────────────────┐\n"
            "│  👑 FACEBOOK DATA MANAGER   │\n"
            "├─────────────────────────────┤\n"
            "│  📊 Please select an option │\n"
            "└─────────────────────────────┘",
            parse_mode="Markdown",
            reply_markup=fb_menu
        )
        return FB_SELECTING_OPTION
    
    if text == "🔐 Use Global":
        pwd = context.user_data.get('fb_global_password')
        if not pwd:
            await update.message.reply_text(
                "┌─────────────────────────────┐\n"
                "│  ❌ NO GLOBAL PASSWORD      │\n"
                "├─────────────────────────────┤\n"
                "│  Please set a password      │\n"
                "│  first using 'Set Password' │\n"
                "└─────────────────────────────┘",
                parse_mode="Markdown",
                reply_markup=BACK_FB_MENU
            )
            return FB_WAITING_PASSWORD
        context.user_data['fb_current_password'] = pwd
        await update.message.reply_text(
            "┌─────────────────────────────┐\n"
            "│  🍪 ENTER COOKIES           │\n"
            "├─────────────────────────────┤\n"
            "│  Please enter your Facebook │\n"
            "│  account cookies            │\n"
            "│  (or type 'skip')           │\n"
            "├─────────────────────────────┤\n"
            "│  🔙 BACK to cancel          │\n"
            "└─────────────────────────────┘",
            parse_mode="Markdown",
            reply_markup=BACK_FB_MENU
        )
        return FB_WAITING_COOKIES
    
    elif text == "✏️ New Password":
        await update.message.reply_text(
            "┌─────────────────────────────┐\n"
            "│  🔑 ENTER NEW PASSWORD      │\n"
            "├─────────────────────────────┤\n"
            "│  Please enter your new      │\n"
            "│  password for this account  │\n"
            "├─────────────────────────────┤\n"
            "│  🔙 BACK to cancel          │\n"
            "└─────────────────────────────┘",
            parse_mode="Markdown",
            reply_markup=BACK_FB_MENU
        )
        return FB_WAITING_PASSWORD_INPUT
    
    elif text == "⏭️ Skip":
        context.user_data['fb_current_password'] = ""
        await update.message.reply_text(
            "┌─────────────────────────────┐\n"
            "│  🍪 ENTER COOKIES           │\n"
            "├─────────────────────────────┤\n"
            "│  Please enter your Facebook │\n"
            "│  account cookies            │\n"
            "│  (or type 'skip')           │\n"
            "├─────────────────────────────┤\n"
            "│  🔙 BACK to cancel          │\n"
            "└─────────────────────────────┘",
            parse_mode="Markdown",
            reply_markup=BACK_FB_MENU
        )
        return FB_WAITING_COOKIES
    
    else:
        await update.message.reply_text(
            "┌─────────────────────────────┐\n"
            "│  ❌ INVALID OPTION          │\n"
            "├─────────────────────────────┤\n"
            "│  Please use the buttons     │\n"
            "└─────────────────────────────┘",
            parse_mode="Markdown"
        )
        return FB_WAITING_PASSWORD

async def fb_waiting_password_input(update: Update, context):
    text = update.message.text
    
    if text == "🔙 BACK":
        user_id = context.user_data.get('fb_user_id')
        count_1000 = get_fb_count(user_id, "1000X")
        count_6155 = get_fb_count(user_id, "6155X")
        
        fb_menu = ReplyKeyboardMarkup([
            [KeyboardButton(f"📁 1000X ({count_1000})", style="success"), KeyboardButton(f"📁 6155X ({count_6155})", style="primary")],
            [KeyboardButton("🔒 Set Password", style="danger"), KeyboardButton("📥 Download", style="success")],
            [KeyboardButton("⚙️ Settings", style="primary")],
            [KeyboardButton("🔙 BACK TO MAIN MENU", style="danger")]
        ], resize_keyboard=True)
        
        await update.message.reply_text(
            "┌─────────────────────────────┐\n"
            "│  👑 FACEBOOK DATA MANAGER   │\n"
            "├─────────────────────────────┤\n"
            "│  📊 Please select an option │\n"
            "└─────────────────────────────┘",
            parse_mode="Markdown",
            reply_markup=fb_menu
        )
        return FB_SELECTING_OPTION
    
    context.user_data['fb_current_password'] = text
    
    await update.message.reply_text(
        "┌─────────────────────────────┐\n"
        "│  🍪 ENTER COOKIES           │\n"
        "├─────────────────────────────┤\n"
        "│  Please enter your Facebook │\n"
        "│  account cookies            │\n"
        "│  (or type 'skip')           │\n"
        "├─────────────────────────────┤\n"
        "│  🔙 BACK to cancel          │\n"
        "└─────────────────────────────┘",
        parse_mode="Markdown",
        reply_markup=BACK_FB_MENU
    )
    return FB_WAITING_COOKIES

async def fb_waiting_cookies(update: Update, context):
    text = update.message.text
    
    if text == "🔙 BACK":
        user_id = context.user_data.get('fb_user_id')
        count_1000 = get_fb_count(user_id, "1000X")
        count_6155 = get_fb_count(user_id, "6155X")
        
        fb_menu = ReplyKeyboardMarkup([
            [KeyboardButton(f"📁 1000X ({count_1000})", style="success"), KeyboardButton(f"📁 6155X ({count_6155})", style="primary")],
            [KeyboardButton("🔒 Set Password", style="danger"), KeyboardButton("📥 Download", style="success")],
            [KeyboardButton("⚙️ Settings", style="primary")],
            [KeyboardButton("🔙 BACK TO MAIN MENU", style="danger")]
        ], resize_keyboard=True)
        
        await update.message.reply_text(
            "┌─────────────────────────────┐\n"
            "│  👑 FACEBOOK DATA MANAGER   │\n"
            "├─────────────────────────────┤\n"
            "│  📊 Please select an option │\n"
            "└─────────────────────────────┘",
            parse_mode="Markdown",
            reply_markup=fb_menu
        )
        return FB_SELECTING_OPTION
    
    cookies = "" if text.lower() == "skip" else text
    
    context.user_data['fb_temp_uid'] = context.user_data.get('fb_current_uid')
    context.user_data['fb_temp_password'] = context.user_data.get('fb_current_password', '')
    context.user_data['fb_temp_cookies'] = cookies
    context.user_data['fb_temp_option'] = context.user_data.get('fb_current_option')
    
    confirm_menu = ReplyKeyboardMarkup([
        [KeyboardButton("✅ Yes", style="success"), KeyboardButton("❌ No", style="danger")],
        [KeyboardButton("🔙 BACK", style="danger")]
    ], resize_keyboard=True)
    
    # Truncate long values for display
    pwd_display = context.user_data['fb_temp_password'][:15] if context.user_data['fb_temp_password'] else "Skip"
    cookies_display = cookies[:30] + "..." if len(cookies) > 30 else (cookies if cookies else "Skip")
    
    await update.message.reply_text(
        f"╔══════════════════════════════════╗\n"
        f"║        ✨ PREVIEW DATA ✨         ║\n"
        f"╠══════════════════════════════════╣\n"
        f"║  🆔 UID: `{context.user_data['fb_temp_uid']}`\n"
        f"║  🔐 Password: `{pwd_display}`\n"
        f"║  🍪 Cookies: `{cookies_display}`\n"
        f"╠══════════════════════════════════╣\n"
        f"║  💾 Save this data?              ║\n"
        f"╚══════════════════════════════════╝",
        parse_mode="Markdown",
        reply_markup=confirm_menu
    )
    return FB_CONFIRM_SAVE

async def fb_confirm_save(update: Update, context):
    text = update.message.text
    user_id = context.user_data.get('fb_user_id')
    
    if text == "🔙 BACK":
        count_1000 = get_fb_count(user_id, "1000X")
        count_6155 = get_fb_count(user_id, "6155X")
        
        fb_menu = ReplyKeyboardMarkup([
            [KeyboardButton(f"📁 1000X ({count_1000})", style="success"), KeyboardButton(f"📁 6155X ({count_6155})", style="primary")],
            [KeyboardButton("🔒 Set Password", style="danger"), KeyboardButton("📥 Download", style="success")],
            [KeyboardButton("⚙️ Settings", style="primary")],
            [KeyboardButton("🔙 BACK TO MAIN MENU", style="danger")]
        ], resize_keyboard=True)
        
        await update.message.reply_text(
            "┌─────────────────────────────┐\n"
            "│  👑 FACEBOOK DATA MANAGER   │\n"
            "├─────────────────────────────┤\n"
            "│  📊 Please select an option │\n"
            "└─────────────────────────────┘",
            parse_mode="Markdown",
            reply_markup=fb_menu
        )
        return FB_SELECTING_OPTION
    
    if text == "✅ Yes":
        uid = context.user_data.get('fb_temp_uid')
        pwd = context.user_data.get('fb_temp_password', '')
        ck = context.user_data.get('fb_temp_cookies', '')
        opt = context.user_data.get('fb_temp_option')
        
        path = get_fb_file_path(user_id, opt)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        
        row_num = None
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == uid:
                row_num = row
                break
        
        if row_num:
            ws.cell(row=row_num, column=2, value=pwd)
            ws.cell(row=row_num, column=3, value=ck)
            action = "UPDATED"
        else:
            new_row = ws.max_row + 1
            ws.cell(row=new_row, column=1, value=uid)
            ws.cell(row=new_row, column=2, value=pwd)
            ws.cell(row=new_row, column=3, value=ck)
            action = "ADDED"
        
        wb.save(path)
        total = ws.max_row
        
        await update.message.reply_text(
            f"╔══════════════════════════════╗\n"
            f"║  ✅ DATA {action} SUCCESSFULLY!   ║\n"
            f"╠══════════════════════════════╣\n"
            f"║  📁 File: {opt}\n"
            f"║  📊 Total: {total} entries\n"
            f"╚══════════════════════════════╝",
            parse_mode="Markdown",
            reply_markup=MAIN_MENU
        )
        
    else:
        await update.message.reply_text(
            "┌─────────────────────────────┐\n"
            "│  ❌ OPERATION CANCELLED     │\n"
            "└─────────────────────────────┘",
            parse_mode="Markdown",
            reply_markup=MAIN_MENU
        )
    
    return ConversationHandler.END

# ==================== মেইন ====================
def main():
    TOKEN = "8583436962:AAGMDPn-yeJJrQLWy53Sc4H6NUXvVtLiP2k"
    
    app = Application.builder().token(TOKEN).build()
    
    # ফাইল রিনেমার কনভার্সেশন
    rename_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^📁 FILE CHANGE$"), handle_menu)],
        states={
            WAITING_FOR_RENAME_FILE: [MessageHandler(filters.ALL, handle_rename_file)],
            WAITING_FOR_NEW_FILENAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_new_filename)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )
    
    # ফেসবুক কুকিজ কনভার্সেশন
    fb_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^📱 FACEBOOK COOKIES$"), handle_menu)],
        states={
            FB_SELECTING_OPTION: [MessageHandler(filters.TEXT & ~filters.COMMAND, fb_select_option)],
            FB_SETTING_PASSWORD: [MessageHandler(filters.TEXT & ~filters.COMMAND, fb_set_password)],
            FB_WAITING_UID: [MessageHandler(filters.TEXT & ~filters.COMMAND, fb_waiting_uid)],
            FB_WAITING_PASSWORD: [MessageHandler(filters.TEXT & ~filters.COMMAND, fb_waiting_password)],
            FB_WAITING_PASSWORD_INPUT: [MessageHandler(filters.TEXT & ~filters.COMMAND, fb_waiting_password_input)],
            FB_WAITING_COOKIES: [MessageHandler(filters.TEXT & ~filters.COMMAND, fb_waiting_cookies)],
            FB_CONFIRM_SAVE: [MessageHandler(filters.TEXT & ~filters.COMMAND, fb_confirm_save)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )
    
    app.add_handler(CommandHandler("start", start))
    app.add_handler(rename_handler)
    app.add_handler(fb_handler)
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_menu))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_number_file))
    app.add_handler(CallbackQueryHandler(toggle_random_callback, pattern='toggle_random'))
    app.add_handler(CallbackQueryHandler(num_convert_callback, pattern='num_convert'))
    app.add_handler(CallbackQueryHandler(num_convert_plus_callback, pattern='num_convert_plus'))
    app.add_handler(CallbackQueryHandler(num_reset_callback, pattern='num_reset'))
    
    print(f"✅ Bot Started! Supporting {len(COUNTRY_DATA)} countries!")
    app.run_polling()

if __name__ == "__main__":
    main()