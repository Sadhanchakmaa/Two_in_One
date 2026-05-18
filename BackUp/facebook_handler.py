# facebook_handler.py
import os
import openpyxl
from telegram import Update, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import ConversationHandler, MessageHandler, filters, CommandHandler, ContextTypes

from config import MAIN_MENU, BACK_FB_MENU, BASE_DIR

# স্টেটস
FB_SELECTING_OPTION = 10
FB_SETTING_PASSWORD = 11
FB_WAITING_UID = 12
FB_WAITING_PASSWORD = 13
FB_WAITING_PASSWORD_INPUT = 14
FB_WAITING_COOKIES = 15
FB_CONFIRM_SAVE = 16


def get_fb_file_path(user_id, option):
    if option == "1000X":
        return os.path.join(BASE_DIR, f"{user_id}_1000X.xlsx")
    else:
        return os.path.join(BASE_DIR, f"{user_id}_6155X.xlsx")


def get_fb_count(user_id, option):
    path = get_fb_file_path(user_id, option)
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        return wb.active.max_row
    return 0


def validate_fb_uid(uid, option):
    if option == "1000X":
        return uid.startswith("1000")
    else:
        return uid.startswith(("6155", "6156", "6157", "6158"))


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text("✨ Cancelled! Returning to Main Menu 🏠", reply_markup=MAIN_MENU)
    return ConversationHandler.END


async def fb_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    context.user_data['fb_user_id'] = user_id
    count_1000 = get_fb_count(user_id, "1000X")
    count_6155 = get_fb_count(user_id, "6155X")
    
    fb_menu = ReplyKeyboardMarkup([
        [KeyboardButton(f"📁 1000X ({count_1000})", style="success"), KeyboardButton(f"📁 6155X ({count_6155})", style="primary")],
        [KeyboardButton("🔒 Set Password", style="danger"), KeyboardButton("📥 Download", style="success")],
        [KeyboardButton("⚙️ Settings", style="primary")],
        [KeyboardButton("🔙 BACK TO MAIN MENU", style="danger")]
    ], resize_keyboard=True)
    
    await update.message.reply_text(
        "✨ *━━━━━━━━━━━━━━━━━━━━━━━━━* ✨\n👑 *VIP FACEBOOK DATA MANAGER* 👑\n✨ *━━━━━━━━━━━━━━━━━━━━━━━━━* ✨\n\n"
        "💎 *Welcome to Team X Bot* 💎\n🎯 *System Ready*\n\n📊 *Please select an option below:*",
        parse_mode="Markdown", reply_markup=fb_menu
    )
    return FB_SELECTING_OPTION


async def fb_select_option(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    user_id = context.user_data.get('fb_user_id')
    
    if text == "🔙 BACK":
        count_1000 = get_fb_count(user_id, "1000X")
        count_6155 = get_fb_count(user_id, "6155X")
        fb_menu = ReplyKeyboardMarkup([
            [KeyboardButton(f"📁 1000X ({count_1000})", style="success"), KeyboardButton(f"📁 6155X ({count_6155})", style="primary")],
            [KeyboardButton("🔒 Set Password", style="danger"), KeyboardButton("📥 Download", style="success")],
            [KeyboardButton("⚙️ Settings", style="primary")],
            [KeyboardButton("🔙 BACK TO MAIN MENU", style="danger")]
        ], resize_keyboard=True)
        await update.message.reply_text("┌─────────────────────────────┐\n│  👑 FACEBOOK DATA MANAGER   │\n├─────────────────────────────┤\n│  📊 Please select an option │\n└─────────────────────────────┘", parse_mode="Markdown", reply_markup=fb_menu)
        return FB_SELECTING_OPTION
    
    if text == "🔙 BACK TO MAIN MENU":
        await update.message.reply_text("┌─────────────────────┐\n│  🏠 BACK TO MENU    │\n└─────────────────────┘", parse_mode="Markdown", reply_markup=MAIN_MENU)
        return ConversationHandler.END
    
    # ========== শুধু মেইন FB মেনুর 1000X বাটন (UID যোগ করার জন্য) ==========
    # "📁 1000X (5)" এই ফরম্যাট চেক করবে
    if text.startswith("📁 1000X") and "File" not in text:
        context.user_data['fb_current_option'] = "1000X"
        path = get_fb_file_path(user_id, "1000X")
        if not os.path.exists(path):
            wb = openpyxl.Workbook()
            wb.save(path)
        await update.message.reply_text(
            "┌─────────────────────────────┐\n│  📱 ENTER FB ACCOUNT UID    │\n├─────────────────────────────┤\n"
            "│  📝 Format:                 │\n│  Must start with 1000       │\n│  Example: 1000123456789     │\n├─────────────────────────────┤\n│  🔙 BACK to cancel          │\n└─────────────────────────────┘",
            parse_mode="Markdown", reply_markup=BACK_FB_MENU
        )
        return FB_WAITING_UID
    
    # ========== শুধু মেইন FB মেনুর 6155X বাটন (UID যোগ করার জন্য) ==========
    if text.startswith("📁 6155X") and "File" not in text:
        context.user_data['fb_current_option'] = "6155X"
        path = get_fb_file_path(user_id, "6155X")
        if not os.path.exists(path):
            wb = openpyxl.Workbook()
            wb.save(path)
        await update.message.reply_text(
            "┌─────────────────────────────┐\n│  📱 ENTER FB ACCOUNT UID    │\n├─────────────────────────────┤\n"
            "│  📝 Format:                 │\n│  Must start with:           │\n│  6155, 6156, 6157, 6158    │\n│  Example: 6155123456789     │\n├─────────────────────────────┤\n│  🔙 BACK to cancel          │\n└─────────────────────────────┘",
            parse_mode="Markdown", reply_markup=BACK_FB_MENU
        )
        return FB_WAITING_UID
    
    elif text == "🔒 Set Password":
        await update.message.reply_text(
            "┌─────────────────────────────┐\n│  🔐 SET GLOBAL PASSWORD     │\n├─────────────────────────────┤\n"
            "│  Enter your password        │\n│  Type 'skip' to continue    │\n│  without password           │\n├─────────────────────────────┤\n│  🔙 BACK to cancel          │\n└─────────────────────────────┘",
            parse_mode="Markdown", reply_markup=BACK_FB_MENU
        )
        return FB_SETTING_PASSWORD
    
    elif text == "⚙️ Settings":
        settings_menu = ReplyKeyboardMarkup([
            [KeyboardButton("🗑️ Reset Data", style="primary")],
            [KeyboardButton("💾 Save Data", style="success")],
            [KeyboardButton("🔙 BACK", style="danger")]
        ], resize_keyboard=True)
        await update.message.reply_text("┌─────────────────────────────┐\n│  ⚙️ SETTINGS PANEL          │\n├─────────────────────────────┤\n│  Configure your preferences │\n└─────────────────────────────┘", parse_mode="Markdown", reply_markup=settings_menu)
        return FB_SELECTING_OPTION
    
    elif text == "🗑️ Reset Data":
        for opt in ["1000X", "6155X"]:
            path = get_fb_file_path(user_id, opt)
            if os.path.exists(path):
                os.remove(path)
        
        count_1000 = get_fb_count(user_id, "1000X")
        count_6155 = get_fb_count(user_id, "6155X")
        fb_menu = ReplyKeyboardMarkup([
            [KeyboardButton(f"📁 1000X ({count_1000})", style="success"), KeyboardButton(f"📁 6155X ({count_6155})", style="primary")],
            [KeyboardButton("🔒 Set Password", style="danger"), KeyboardButton("📥 Download", style="success")],
            [KeyboardButton("⚙️ Settings", style="primary")],
            [KeyboardButton("🔙 BACK TO MAIN MENU", style="danger")]
        ], resize_keyboard=True)
        
        await update.message.reply_text(
            "┌─────────────────────────────┐\n│  ✅ DATA CLEARED            │\n│  All files have been deleted│\n├─────────────────────────────┤\n│  👑 FACEBOOK DATA MANAGER   │\n│  📊 Please select an option │\n└─────────────────────────────┘", 
            parse_mode="Markdown", 
            reply_markup=fb_menu
        )
        return FB_SELECTING_OPTION
    
    elif text == "💾 Save Data":
        count_1000 = get_fb_count(user_id, "1000X")
        count_6155 = get_fb_count(user_id, "6155X")
        fb_menu = ReplyKeyboardMarkup([
            [KeyboardButton(f"📁 1000X ({count_1000})", style="success"), KeyboardButton(f"📁 6155X ({count_6155})", style="primary")],
            [KeyboardButton("🔒 Set Password", style="danger"), KeyboardButton("📥 Download", style="success")],
            [KeyboardButton("⚙️ Settings", style="primary")],
            [KeyboardButton("🔙 BACK TO MAIN MENU", style="danger")]
        ], resize_keyboard=True)
        
        await update.message.reply_text(
            "┌─────────────────────────────┐\n│  💾 DATA SAVED              │\n│  All changes saved          │\n├─────────────────────────────┤\n│  👑 FACEBOOK DATA MANAGER   │\n│  📊 Please select an option │\n└─────────────────────────────┘", 
            parse_mode="Markdown", 
            reply_markup=fb_menu
        )
        return FB_SELECTING_OPTION
    
    elif text == "📥 Download":
        download_menu = ReplyKeyboardMarkup([
            [KeyboardButton("📁 1000X File", style="success"), KeyboardButton("📁 6155X File", style="primary")],
            [KeyboardButton("📁 Both Files", style="success")],
            [KeyboardButton("🔙 BACK", style="danger")]
        ], resize_keyboard=True)
        await update.message.reply_text("┌─────────────────────────────┐\n│  📥 DOWNLOAD CENTER         │\n├─────────────────────────────┤\n│  Select file to download:   │\n└─────────────────────────────┘", parse_mode="Markdown", reply_markup=download_menu)
        return FB_SELECTING_OPTION
    
    elif text == "📁 1000X File":
        return await fb_download_file(update, context, "1000X")
    
    elif text == "📁 6155X File":
        return await fb_download_file(update, context, "6155X")
    
    elif text == "📁 Both Files":
        return await fb_download_both(update, context)
    
    else:
        await update.message.reply_text("┌─────────────────────────────┐\n│  ❌ INVALID OPTION          │\n├─────────────────────────────┤\n│  Please use the buttons     │\n└─────────────────────────────┘", parse_mode="Markdown")
        return FB_SELECTING_OPTION


async def fb_download_file(update: Update, context: ContextTypes.DEFAULT_TYPE, option):
    user_id = context.user_data.get('fb_user_id')
    path = get_fb_file_path(user_id, option)
    
    if not os.path.exists(path):
        await update.message.reply_text(f"┌─────────────────────────────┐\n│  ❌ NO DATA FOUND           │\n├─────────────────────────────┤\n│  📁 {option} has no data    │\n├─────────────────────────────┤\n│  📤 Add data first          │\n└─────────────────────────────┘", parse_mode="Markdown")
        return FB_SELECTING_OPTION
    
    wb = openpyxl.load_workbook(path)
    total = wb.active.max_row
    
    with open(path, "rb") as f:
        await update.message.reply_document(document=f, filename=f"{option}.xlsx",
            caption=f"╔══════════════════════════════╗\n║  ✅ FILE READY FOR DOWNLOAD  ║\n╠══════════════════════════════╣\n"
                    f"║  📁 File: {option:<16}║\n║  📊 Total: {total} entries             ║\n╚══════════════════════════════╝")
    
    return FB_SELECTING_OPTION


async def fb_download_both(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = context.user_data.get('fb_user_id')
    sent_count = 0
    
    for opt in ["1000X", "6155X"]:
        path = get_fb_file_path(user_id, opt)
        if os.path.exists(path):
            wb = openpyxl.load_workbook(path)
            total = wb.active.max_row
            with open(path, "rb") as f:
                await update.message.reply_document(document=f, filename=f"{opt}.xlsx",
                    caption=f"╔════════════════════════╗\n║  📁 {opt:<12}║\n║  📊 {total} entries        ║\n╚════════════════════════╝")
            sent_count += 1
        else:
            await update.message.reply_text(f"┌─────────────────────────┐\n│  ⚠️ NO DATA FOR {opt:<9}│\n└─────────────────────────┘", parse_mode="Markdown")
    
    if sent_count == 0:
        await update.message.reply_text("┌─────────────────────────────┐\n│  ❌ NO FILES AVAILABLE      │\n├─────────────────────────────┤\n│  No data found in both      │\n│  1000X and 6155X folders    │\n└─────────────────────────────┘", parse_mode="Markdown")
    
    return FB_SELECTING_OPTION


async def fb_set_password(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    user_id = context.user_data.get('fb_user_id')
    
    if text == "🔙 BACK":
        count_1000 = get_fb_count(user_id, "1000X")
        count_6155 = get_fb_count(user_id, "6155X")
        fb_menu = ReplyKeyboardMarkup([
            [KeyboardButton(f"📁 1000X ({count_1000})", style="success"), KeyboardButton(f"📁 6155X ({count_6155})", style="primary")],
            [KeyboardButton("🔒 Set Password", style="danger"), KeyboardButton("📥 Download", style="success")],
            [KeyboardButton("⚙️ Settings", style="primary")],
            [KeyboardButton("🔙 BACK TO MAIN MENU", style="danger")]
        ], resize_keyboard=True)
        await update.message.reply_text("┌─────────────────────────────┐\n│  👑 FACEBOOK DATA MANAGER   │\n├─────────────────────────────┤\n│  📊 Please select an option │\n└─────────────────────────────┘", parse_mode="Markdown", reply_markup=fb_menu)
        return FB_SELECTING_OPTION
    
    if text.lower() == "skip":
        context.user_data['fb_global_password'] = None
        await update.message.reply_text("┌─────────────────────────────┐\n│  ⏭️ PASSWORD SKIPPED        │\n├─────────────────────────────┤\n│  No password will be set    │\n└─────────────────────────────┘", parse_mode="Markdown")
    else:
        context.user_data['fb_global_password'] = text
        updated_count = 0
        for opt in ["1000X", "6155X"]:
            path = get_fb_file_path(user_id, opt)
            if os.path.exists(path):
                wb = openpyxl.load_workbook(path)
                ws = wb.active
                for row in range(1, ws.max_row + 1):
                    ws.cell(row=row, column=2, value=text)
                wb.save(path)
                updated_count += 1
        await update.message.reply_text(f"╔══════════════════════════════╗\n║  🔐 PASSWORD SET SUCCESSFULLY ║\n╠══════════════════════════════╣\n║  📁 Files updated: {updated_count}           ║\n║  🔒 Password: `{text[:10]}...`    ║\n╚══════════════════════════════╝", parse_mode="Markdown")
    
    count_1000 = get_fb_count(user_id, "1000X")
    count_6155 = get_fb_count(user_id, "6155X")
    fb_menu = ReplyKeyboardMarkup([
        [KeyboardButton(f"📁 1000X ({count_1000})", style="success"), KeyboardButton(f"📁 6155X ({count_6155})", style="primary")],
        [KeyboardButton("🔒 Set Password", style="danger"), KeyboardButton("📥 Download", style="success")],
        [KeyboardButton("⚙️ Settings", style="primary")],
        [KeyboardButton("🔙 BACK TO MAIN MENU", style="danger")]
    ], resize_keyboard=True)
    await update.message.reply_text("┌─────────────────────────────┐\n│  👑 FACEBOOK DATA MANAGER   │\n├─────────────────────────────┤\n│  📊 Please select an option │\n└─────────────────────────────┘", parse_mode="Markdown", reply_markup=fb_menu)
    return FB_SELECTING_OPTION


async def fb_waiting_uid(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    current_option = context.user_data.get('fb_current_option')
    
    if text == "🔙 BACK":
        user_id = context.user_data.get('fb_user_id')
        count_1000 = get_fb_count(user_id, "1000X")
        count_6155 = get_fb_count(user_id, "6155X")
        fb_menu = ReplyKeyboardMarkup([
            [KeyboardButton(f"📁 1000X ({count_1000})", style="success"), KeyboardButton(f"📁 6155X ({count_6155})", style="primary")],
            [KeyboardButton("🔒 Set Password", style="danger"), KeyboardButton("📥 Download", style="success")],
            [KeyboardButton("⚙️ Settings", style="primary")],
            [KeyboardButton("🔙 BACK TO MAIN MENU", style="danger")]
        ], resize_keyboard=True)
        await update.message.reply_text("┌─────────────────────────────┐\n│  👑 FACEBOOK DATA MANAGER   │\n├─────────────────────────────┤\n│  📊 Please select an option │\n└─────────────────────────────┘", parse_mode="Markdown", reply_markup=fb_menu)
        return FB_SELECTING_OPTION
    
    if not validate_fb_uid(text, current_option):
        await update.message.reply_text(f"┌─────────────────────────────────┐\n│  ❌ INVALID UID FOR {current_option}  │\n├─────────────────────────────────┤\n│  📝 Please enter a valid UID:    │\n│  🔙 BACK to cancel              │\n└─────────────────────────────────┘", parse_mode="Markdown", reply_markup=BACK_FB_MENU)
        return FB_WAITING_UID
    
    context.user_data['fb_current_uid'] = text
    
    pwd_menu = ReplyKeyboardMarkup([
        [KeyboardButton("🔐 Use Global", style="success"), KeyboardButton("✏️ New Password", style="primary")],
        [KeyboardButton("⏭️ Skip", style="success")],
        [KeyboardButton("🔙 BACK", style="danger")]
    ], resize_keyboard=True)
    
    await update.message.reply_text("┌─────────────────────────────┐\n│  🔐 PASSWORD OPTIONS        │\n├─────────────────────────────┤\n│  How would you like to set  │\n│  the password?              │\n└─────────────────────────────┘", parse_mode="Markdown", reply_markup=pwd_menu)
    return FB_WAITING_PASSWORD


async def fb_waiting_password(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    
    if text == "🔙 BACK":
        user_id = context.user_data.get('fb_user_id')
        count_1000 = get_fb_count(user_id, "1000X")
        count_6155 = get_fb_count(user_id, "6155X")
        fb_menu = ReplyKeyboardMarkup([
            [KeyboardButton(f"📁 1000X ({count_1000})", style="success"), KeyboardButton(f"📁 6155X ({count_6155})", style="primary")],
            [KeyboardButton("🔒 Set Password", style="danger"), KeyboardButton("📥 Download", style="success")],
            [KeyboardButton("⚙️ Settings", style="primary")],
            [KeyboardButton("🔙 BACK TO MAIN MENU", style="danger")]
        ], resize_keyboard=True)
        await update.message.reply_text("┌─────────────────────────────┐\n│  👑 FACEBOOK DATA MANAGER   │\n├─────────────────────────────┤\n│  📊 Please select an option │\n└─────────────────────────────┘", parse_mode="Markdown", reply_markup=fb_menu)
        return FB_SELECTING_OPTION
    
    if text == "🔐 Use Global":
        pwd = context.user_data.get('fb_global_password')
        if not pwd:
            await update.message.reply_text("┌─────────────────────────────┐\n│  ❌ NO GLOBAL PASSWORD      │\n├─────────────────────────────┤\n│  Please set a password      │\n│  first using 'Set Password' │\n└─────────────────────────────┘", parse_mode="Markdown", reply_markup=BACK_FB_MENU)
            return FB_WAITING_PASSWORD
        context.user_data['fb_current_password'] = pwd
        await update.message.reply_text("┌─────────────────────────────┐\n│  🍪 ENTER COOKIES           │\n├─────────────────────────────┤\n│  Please enter your Facebook │\n│  account cookies            │\n│  (or type 'skip')           │\n├─────────────────────────────┤\n│  🔙 BACK to cancel          │\n└─────────────────────────────┘", parse_mode="Markdown", reply_markup=BACK_FB_MENU)
        return FB_WAITING_COOKIES
    
    elif text == "✏️ New Password":
        await update.message.reply_text("┌─────────────────────────────┐\n│  🔑 ENTER NEW PASSWORD      │\n├─────────────────────────────┤\n│  Please enter your new      │\n│  password for this account  │\n├─────────────────────────────┤\n│  🔙 BACK to cancel          │\n└─────────────────────────────┘", parse_mode="Markdown", reply_markup=BACK_FB_MENU)
        return FB_WAITING_PASSWORD_INPUT
    
    elif text == "⏭️ Skip":
        context.user_data['fb_current_password'] = ""
        await update.message.reply_text("┌─────────────────────────────┐\n│  🍪 ENTER COOKIES           │\n├─────────────────────────────┤\n│  Please enter your Facebook │\n│  account cookies            │\n│  (or type 'skip')           │\n├─────────────────────────────┤\n│  🔙 BACK to cancel          │\n└─────────────────────────────┘", parse_mode="Markdown", reply_markup=BACK_FB_MENU)
        return FB_WAITING_COOKIES
    
    else:
        await update.message.reply_text("┌─────────────────────────────┐\n│  ❌ INVALID OPTION          │\n├─────────────────────────────┤\n│  Please use the buttons     │\n└─────────────────────────────┘", parse_mode="Markdown")
        return FB_WAITING_PASSWORD


async def fb_waiting_password_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    
    if text == "🔙 BACK":
        user_id = context.user_data.get('fb_user_id')
        count_1000 = get_fb_count(user_id, "1000X")
        count_6155 = get_fb_count(user_id, "6155X")
        fb_menu = ReplyKeyboardMarkup([
            [KeyboardButton(f"📁 1000X ({count_1000})", style="success"), KeyboardButton(f"📁 6155X ({count_6155})", style="primary")],
            [KeyboardButton("🔒 Set Password", style="danger"), KeyboardButton("📥 Download", style="success")],
            [KeyboardButton("⚙️ Settings", style="primary")],
            [KeyboardButton("🔙 BACK TO MAIN MENU", style="danger")]
        ], resize_keyboard=True)
        await update.message.reply_text("┌─────────────────────────────┐\n│  👑 FACEBOOK DATA MANAGER   │\n├─────────────────────────────┤\n│  📊 Please select an option │\n└─────────────────────────────┘", parse_mode="Markdown", reply_markup=fb_menu)
        return FB_SELECTING_OPTION
    
    context.user_data['fb_current_password'] = text
    
    await update.message.reply_text("┌─────────────────────────────┐\n│  🍪 ENTER COOKIES           │\n├─────────────────────────────┤\n│  Please enter your Facebook │\n│  account cookies            │\n│  (or type 'skip')           │\n├─────────────────────────────┤\n│  🔙 BACK to cancel          │\n└─────────────────────────────┘", parse_mode="Markdown", reply_markup=BACK_FB_MENU)
    return FB_WAITING_COOKIES


async def fb_waiting_cookies(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    
    if text == "🔙 BACK":
        user_id = context.user_data.get('fb_user_id')
        count_1000 = get_fb_count(user_id, "1000X")
        count_6155 = get_fb_count(user_id, "6155X")
        fb_menu = ReplyKeyboardMarkup([
            [KeyboardButton(f"📁 1000X ({count_1000})", style="success"), KeyboardButton(f"📁 6155X ({count_6155})", style="primary")],
            [KeyboardButton("🔒 Set Password", style="danger"), KeyboardButton("📥 Download", style="success")],
            [KeyboardButton("⚙️ Settings", style="primary")],
            [KeyboardButton("🔙 BACK TO MAIN MENU", style="danger")]
        ], resize_keyboard=True)
        await update.message.reply_text("┌─────────────────────────────┐\n│  👑 FACEBOOK DATA MANAGER   │\n├─────────────────────────────┤\n│  📊 Please select an option │\n└─────────────────────────────┘", parse_mode="Markdown", reply_markup=fb_menu)
        return FB_SELECTING_OPTION
    
    cookies = "" if text.lower() == "skip" else text
    
    context.user_data['fb_temp_uid'] = context.user_data.get('fb_current_uid')
    context.user_data['fb_temp_password'] = context.user_data.get('fb_current_password', '')
    context.user_data['fb_temp_cookies'] = cookies
    context.user_data['fb_temp_option'] = context.user_data.get('fb_current_option')
    
    confirm_menu = ReplyKeyboardMarkup([
        [KeyboardButton("✅ Yes", style="success"), KeyboardButton("❌ No", style="danger")],
        [KeyboardButton("🔙 BACK", style="danger")]
    ], resize_keyboard=True)
    
    pwd_display = context.user_data['fb_temp_password'][:15] if context.user_data['fb_temp_password'] else "Skip"
    cookies_display = cookies[:30] + "..." if len(cookies) > 30 else (cookies if cookies else "Skip")
    
    await update.message.reply_text(
        f"╔══════════════════════════════════╗\n║        ✨ PREVIEW DATA ✨         ║\n╠══════════════════════════════════╣\n"
        f"║  🆔 UID: `{context.user_data['fb_temp_uid']}`\n║  🔐 Password: `{pwd_display}`\n║  🍪 Cookies: `{cookies_display}`\n"
        f"╠══════════════════════════════════╣\n║  💾 Save this data?              ║\n╚══════════════════════════════════╝",
        parse_mode="Markdown", reply_markup=confirm_menu
    )
    return FB_CONFIRM_SAVE


async def fb_confirm_save(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    user_id = context.user_data.get('fb_user_id')
    
    if text == "🔙 BACK":
        count_1000 = get_fb_count(user_id, "1000X")
        count_6155 = get_fb_count(user_id, "6155X")
        fb_menu = ReplyKeyboardMarkup([
            [KeyboardButton(f"📁 1000X ({count_1000})", style="success"), KeyboardButton(f"📁 6155X ({count_6155})", style="primary")],
            [KeyboardButton("🔒 Set Password", style="danger"), KeyboardButton("📥 Download", style="success")],
            [KeyboardButton("⚙️ Settings", style="primary")],
            [KeyboardButton("🔙 BACK TO MAIN MENU", style="danger")]
        ], resize_keyboard=True)
        await update.message.reply_text("┌─────────────────────────────┐\n│  👑 FACEBOOK DATA MANAGER   │\n├─────────────────────────────┤\n│  📊 Please select an option │\n└─────────────────────────────┘", parse_mode="Markdown", reply_markup=fb_menu)
        return FB_SELECTING_OPTION
    
    if text == "✅ Yes":
        uid = context.user_data.get('fb_temp_uid')
        pwd = context.user_data.get('fb_temp_password', '')
        ck = context.user_data.get('fb_temp_cookies', '')
        opt = context.user_data.get('fb_temp_option')
        
        path = get_fb_file_path(user_id, opt)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        
        row_num = None
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == uid:
                row_num = row
                break
        
        if row_num:
            ws.cell(row=row_num, column=2, value=pwd)
            ws.cell(row=row_num, column=3, value=ck)
            action = "UPDATED"
        else:
            new_row = ws.max_row + 1
            ws.cell(row=new_row, column=1, value=uid)
            ws.cell(row=new_row, column=2, value=pwd)
            ws.cell(row=new_row, column=3, value=ck)
            action = "ADDED"
        
        wb.save(path)
        total = ws.max_row
        
        # সেভ成功后 Facebook মেনু দেখান
        count_1000 = get_fb_count(user_id, "1000X")
        count_6155 = get_fb_count(user_id, "6155X")
        fb_menu = ReplyKeyboardMarkup([
            [KeyboardButton(f"📁 1000X ({count_1000})", style="success"), KeyboardButton(f"📁 6155X ({count_6155})", style="primary")],
            [KeyboardButton("🔒 Set Password", style="danger"), KeyboardButton("📥 Download", style="success")],
            [KeyboardButton("⚙️ Settings", style="primary")],
            [KeyboardButton("🔙 BACK TO MAIN MENU", style="danger")]
        ], resize_keyboard=True)
        
        await update.message.reply_text(
            f"╔══════════════════════════════╗\n║  ✅ DATA {action} SUCCESSFULLY!   ║\n╠══════════════════════════════╣\n"
            f"║  📁 File: {opt}\n║  📊 Total: {total} entries\n╚══════════════════════════════╝\n\n"
            "┌─────────────────────────────┐\n│  👑 FACEBOOK DATA MANAGER   │\n├─────────────────────────────┤\n│  📊 Please select an option │\n└─────────────────────────────┘",
            parse_mode="Markdown", 
            reply_markup=fb_menu
        )
        return FB_SELECTING_OPTION  # Facebook মেনুতেই থাকবে, ConversationHandler.END না
    
    else:  # ❌ No
        # ক্যান্সেল করার পরও Facebook মেনু দেখান
        count_1000 = get_fb_count(user_id, "1000X")
        count_6155 = get_fb_count(user_id, "6155X")
        fb_menu = ReplyKeyboardMarkup([
            [KeyboardButton(f"📁 1000X ({count_1000})", style="success"), KeyboardButton(f"📁 6155X ({count_6155})", style="primary")],
            [KeyboardButton("🔒 Set Password", style="danger"), KeyboardButton("📥 Download", style="success")],
            [KeyboardButton("⚙️ Settings", style="primary")],
            [KeyboardButton("🔙 BACK TO MAIN MENU", style="danger")]
        ], resize_keyboard=True)
        
        await update.message.reply_text(
            "┌─────────────────────────────┐\n│  ❌ OPERATION CANCELLED     │\n├─────────────────────────────┤\n"
            "│  👑 FACEBOOK DATA MANAGER   │\n│  📊 Please select an option │\n└─────────────────────────────┘", 
            parse_mode="Markdown", 
            reply_markup=fb_menu
        )
        return FB_SELECTING_OPTION  # Facebook মেনুতেই থাকবে


def facebook_conversation_handler():
    return ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^📱 FACEBOOK COOKIES$"), fb_start)],
        states={
            FB_SELECTING_OPTION: [MessageHandler(filters.TEXT & ~filters.COMMAND, fb_select_option)],
            FB_SETTING_PASSWORD: [MessageHandler(filters.TEXT & ~filters.COMMAND, fb_set_password)],
            FB_WAITING_UID: [MessageHandler(filters.TEXT & ~filters.COMMAND, fb_waiting_uid)],
            FB_WAITING_PASSWORD: [MessageHandler(filters.TEXT & ~filters.COMMAND, fb_waiting_password)],
            FB_WAITING_PASSWORD_INPUT: [MessageHandler(filters.TEXT & ~filters.COMMAND, fb_waiting_password_input)],
            FB_WAITING_COOKIES: [MessageHandler(filters.TEXT & ~filters.COMMAND, fb_waiting_cookies)],
            FB_CONFIRM_SAVE: [MessageHandler(filters.TEXT & ~filters.COMMAND, fb_confirm_save)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )